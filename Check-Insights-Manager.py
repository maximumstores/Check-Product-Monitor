import time
import logging
from datetime import datetime
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import requests
import re
from urllib.parse import urlparse, parse_qs
import os
from threading import Lock
import ast
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import telebot
import random
import pytz
import json
from gspread.exceptions import APIError
from bs4 import BeautifulSoup  # Добавлено для парсинга HTML, если потребуется
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
import gspread_formatting as gf
from gspread_formatting import *
from google.oauth2.service_account import Credentials

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class APIRateLimiter:
    """Класс для ограничения количества запросов к API."""
    def __init__(self, max_requests, period):
        self.max_requests = max_requests
        self.period = period
        self.requests = []
        self.lock = Lock()

    def wait(self):
        """Ожидание перед следующим запросом, если достигнут лимит."""
        with self.lock:
            now = time.time()
            # Удаляем устаревшие запросы
            self.requests = [r for r in self.requests if r > now - self.period]
            while len(self.requests) >= self.max_requests:
                next_request_time = self.requests[0] + self.period
                sleep_time = max(next_request_time - now, 0)
                logging.debug(f"Достигнут лимит запросов. Спим {sleep_time:.2f} секунд.")
                time.sleep(sleep_time)
                now = time.time()
                self.requests = [r for r in self.requests if r > now - self.period]
            self.requests.append(now)

# Настройка лимитера на 1 запрос в секунду
api_limiter = APIRateLimiter(max_requests=1, period=1)  # 1 запрос в секунду

def clean_urls(raw_value):
    if isinstance(raw_value, str):
        # Разделяем по переносам строк, запятым и пробелам
        urls = [url.strip() for url in re.split(r'[\n\r,]+', raw_value) if url.strip()]
    else:
        urls = raw_value if isinstance(raw_value, list) else []
    return urls



def retry_scrape(url, config, max_retries=5, initial_wait=2):
    wait_time = initial_wait
    for attempt in range(max_retries):
        try:
            api_limiter.wait()  # Ждем, чтобы не превысить лимит запросов
            product_info = scrape_amazon_product(url, config)
            if product_info:
                return product_info
        except requests.exceptions.RequestException as e:
            if hasattr(e, 'response') and e.response and e.response.status_code == 429:
                logging.warning(f"Ошибка 429. Повторная попытка через {wait_time} секунд...")
                time.sleep(wait_time)
                wait_time *= 2  # Экспоненциальное увеличение времени ожидания
            else:
                logging.error(f"Ошибка запроса для URL {url}: {str(e)}")
                break
    return None

def get_kyiv_time(timezone_str='Europe/Kiev'):
    """Возвращает текущее время в часовом поясе Киева."""
    timezone = pytz.timezone(timezone_str)
    return datetime.now(timezone)


def get_next_slot(current_time, slots, timezone_str='Europe/Kiev'):
    """
    Возвращает ближайший следующий слот из списка slots.

    :param current_time: Текущее время (datetime объект)
    :param slots: Список временных слотов в формате "HH:MM"
    :param timezone_str: Часовой пояс
    :return: datetime объект ближайшего следующего слота
    """
    timezone = pytz.timezone(timezone_str)
    today_slots = sorted(slots)
    for slot in today_slots:
        slot_time = datetime.strptime(slot, "%H:%M").time()
        slot_datetime = timezone.localize(datetime.combine(current_time.date(), slot_time))
        if slot_datetime > current_time:
            return slot_datetime
    # Если все слоты на сегодня уже прошли, возвращаем первый слот на завтра
    first_slot = datetime.strptime(today_slots[0], "%H:%M").time()
    next_day = current_time + timedelta(days=1)
    return timezone.localize(datetime.combine(next_day.date(), first_slot))


def authorize_google_sheets(credentials_file):
    """
    Авторизуется в Google Sheets и возвращает клиентский объект.
    
    :param credentials_file: Путь к файлу учетных данных JSON
    :return: gspread.Client объект
    """
    scopes = [
        'https://www.googleapis.com/auth/spreadsheets',
        'https://www.googleapis.com/auth/drive'
    ]
    
    credentials = Credentials.from_service_account_file(credentials_file, scopes=scopes)
    client = gspread.authorize(credentials)
    logging.info("Успешно авторизовались в Google Sheets")
    return client


def load_config_from_sheets(client, spreadsheet_id):
    """Загрузка конфигурации из Google Sheets."""
    try:
        config_sheet = client.open_by_key(spreadsheet_id).worksheet('Config')
    except gspread.exceptions.WorksheetNotFound:
        logging.error("Лист 'Config' не найден в таблице.")
        raise

    config = {}
    all_records = config_sheet.get_all_records()

    for record in all_records:
        key = record.get('Key', '')
        value = record.get('Value', '')

        # Преобразуем 'Key' в строку и удаляем лишние пробелы
        if isinstance(key, str):
            key = key.strip()
        else:
            key = str(key).strip()

        # Преобразуем 'Value' в строку (если это не строка) и удаляем кавычки
        if isinstance(value, str):
            value = value.strip().strip('"').strip("'")
        else:
            value = str(value).strip()

        if not key:
            logging.warning("Пропущена запись с пустым ключом.")
            continue

        # Обработка URL-полей
        url_keys = [
            'product_urls', '1competitor_urls', '2competitor_urls', 
            '3competitor_urls', 'variation_urls', '1variation_urls', 
            '2variation_urls', '3variation_urls'
        ]

        if key in url_keys:
            # Если значение уже список, используем его
            if isinstance(value, list):
                config[key] = value
            else:
                # Иначе, разбиваем строку на список URL
                config[key] = clean_urls(value)
            logging.debug(f"Загружены URL для '{key}': {config[key]}")
        elif key in ['update_time_hour', 'update_time_minute', 'batch_size']:
            try:
                config[key] = int(value)
                logging.debug(f"Загружено целое число для '{key}': {config[key]}")
            except ValueError:
                logging.error(f"Некорректное целое число для '{key}': {value}. Установлено значение по умолчанию 0.")
                config[key] = 0
        elif key in ['min_acceptable_rating', 'price_change_threshold', 'coupon_threshold']:
            try:
                config[key] = float(value)
                logging.debug(f"Загружено число с плавающей точкой для '{key}': {config[key]}")
            except ValueError:
                logging.error(f"Некорректное число с плавающей точкой для '{key}': {value}. Установлено значение по умолчанию 0.0.")
                config[key] = 0.0
        elif key in ['active_trade_slots', 'analysis_slots']:
            if value:
                # Разделяем слоты по запятым и пробелам
                config[key] = [slot.strip() for slot in re.split(r'[,\s]+', value) if slot.strip()]
                logging.debug(f"Загружены временные слоты для '{key}': {config[key]}")
            else:
                config[key] = []
        else:
            config[key] = value
            logging.debug(f"Загружено значение для '{key}': {config[key]}")

    # Обновление имен конкурентов без лишних пробелов
    competitor_1_name = config.get('competitor_1_name', 'Конкурент 1')
    competitor_2_name = config.get('competitor_2_name', 'Конкурент 2')
    competitor_3_name = config.get('competitor_3_name', 'Конкурент 3')

    if isinstance(competitor_1_name, str):
        competitor_1_name = competitor_1_name.strip()
    if isinstance(competitor_2_name, str):
        competitor_2_name = competitor_2_name.strip()
    if isinstance(competitor_3_name, str):
        competitor_3_name = competitor_3_name.strip()

    config['competitor_names'] = {
        '1': competitor_1_name,
        '2': competitor_2_name,
        '3': competitor_3_name
    }

    logging.info(f"Загруженная конфигурация: {json.dumps(config, indent=2, ensure_ascii=False)}")
    return config


def extract_asin(url):
    """Извлекает ASIN из различных форматов URL Amazon."""
    # Парсим URL
    parsed_url = urlparse(url)
    path = parsed_url.path

    # Регулярное выражение для извлечения ASIN из пути URL
    patterns = [
        r'/dp/([A-Z0-9]{10})',
        r'/gp/product/([A-Z0-9]{10})',
        r'/product/([A-Z0-9]{10})',
        r'/ASIN/([A-Z0-9]{10})',
        r'/gp/aw/d/([A-Z0-9]{10})',
        r'/gp/offer-listing/([A-Z0-9]{10})',
    ]
    for pattern in patterns:
        match = re.search(pattern, path)
        if match:
            asin = match.group(1)
            logging.debug(f"Extracted ASIN {asin} from URL path: {url}")
            return asin
    # Если не удалось найти ASIN в пути, попробуем извлечь из параметров запроса
    query_params = parse_qs(parsed_url.query)
    if 'asin' in query_params:
        asin = query_params['asin'][0]
        logging.debug(f"Extracted ASIN {asin} from query parameters in URL: {url}")
        return asin
    # Если всё ещё не удалось найти ASIN, попробуем найти его в URL целиком
    match = re.search(r'([A-Z0-9]{10})', url)
    if match:
        asin = match.group(1)
        logging.debug(f"Extracted ASIN {asin} from entire URL: {url}")
        return asin
    logging.warning(f"Could not extract ASIN from URL: {url}")
    return 'Not Found'

def is_valid_amazon_url(url):
    """Проверяет, является ли URL корректным Amazon продуктом."""
    parsed_url = urlparse(url)
    return parsed_url.netloc in ['www.amazon.com', 'amazon.com'] and '/dp/' in parsed_url.path
def clean_urls(raw_value):
    """
    Очищает строку URL-адресов, корректно обрабатывает любые разделители,
    включая переносы строк, запятые, пробелы, а также разделители с запятыми и переносами строк.
    """
    if isinstance(raw_value, str):
        # Заменяем переносы строк на запятые, затем разделяем по запятым и пробелам
        urls = [url.strip() for url in re.split(r'[\n\r,]+', raw_value) if url.strip()]
    else:
        urls = raw_value if isinstance(raw_value, list) else []

    return urls

def write_hyperlinks(sheet, urls, start_row, column):
    """
    Записывает гиперссылки в Google Sheets, каждая ссылка в отдельной строке.
    
    :param sheet: Лист Google Sheets
    :param urls: Список URL-адресов для записи
    :param start_row: Начальная строка для записи
    :param column: Колонка для записи (например, 'A', 'B')
    """
    for i, url in enumerate(urls):
        asin = extract_asin(url)  # Извлекаем ASIN для отображения
        if asin != 'Not Found':
            # Создаем гиперссылку для каждой ссылки отдельно
            hyperlink_formula = f'=HYPERLINK("{url}", "{asin}")'
            cell = f'{column}{start_row + i}'  # Каждая ссылка в новой строке
            sheet.update_acell(cell, hyperlink_formula)
        else:
            logging.warning(f"ASIN not found for URL: {url}")



def calculate_final_price(full_price, prime_price, coupon_discount):
    """
    Вычисляет итоговую цену с учётом скидок и купонов.
    Возвращает строку с форматом цены.
    """
    try:
        # Преобразование цен из строк в числа
        def price_to_float(price_str):
            if not price_str or price_str == "Not Found":
                return None
            return float(re.sub(r'[^\d.]', '', price_str))

        full_price_value = price_to_float(full_price)
        prime_price_value = price_to_float(prime_price)
        coupon_discount_value = float(re.sub(r'[^\d.]', '', str(coupon_discount).replace('%', ''))) if coupon_discount and coupon_discount != "Not Found" else 0.0

        # Используем prime_price_value, если доступно, иначе full_price_value
        base_price = prime_price_value or full_price_value
        if base_price is None:
            logging.warning("Base price is not available for final price calculation.")
            return "Not Found"

        # Вычисляем итоговую цену с учётом купона
        discount_amount = base_price * (coupon_discount_value / 100)
        final_price_value = base_price - discount_amount

        # Возвращаем итоговую цену в формате строки
        return f"${final_price_value:.2f}"
    except ValueError as e:
        logging.error(f"Ошибка при расчете цены: {str(e)}")
        return "Not Found"

def calculate_discount_percent(full_price, final_price):
    """Вычисляет процент скидки."""
    try:
        if full_price == "Not Found" or final_price == "Not Found":
            return "Не применимо"
        
        full_price_value = float(re.sub(r'[^\d.]', '', str(full_price)))
        final_price_value = float(re.sub(r'[^\d.]', '', str(final_price)))
        
        if full_price_value == 0:
            return "N/A"
        
        discount_percent_value = (full_price_value - final_price_value) / full_price_value * 100
        return f"{discount_percent_value:.2f}%"
    except ValueError:
        logging.error(f"Ошибка при расчете процента скидки с Full Price: {full_price} и Final Price: {final_price}")
        return "Не применимо"
def extract_price(price_data):
    """
    Извлекает цену из данных продукта.
    
    :param price_data: Данные о цене из JSON-ответа.
    :return: Строка с форматом цены или "Not Found".
    """
    logging.debug(f"Извлечение цены из данных: {price_data}")
    
    if not price_data or price_data == "Not Found":
        logging.warning("Цена не найдена в данных продукта.")
        return "Not Found"
    
    if isinstance(price_data, dict):
        # Возможные ключи для цены
        possible_keys = ['raw', 'display_price', 'value', 'price']
        for key in possible_keys:
            if key in price_data:
                extracted_price = price_data[key]
                logging.debug(f"Найдено '{key}': {extracted_price}")
                if isinstance(extracted_price, (int, float)):
                    return f"${extracted_price:.2f}"
                elif isinstance(extracted_price, str):
                    match = re.search(r'\$?\d+(\.\d+)?', extracted_price)
                    if match:
                        return match.group()
        logging.warning("Цена не удалось извлечь из словаря.")
    elif isinstance(price_data, (int, float)):
        logging.debug(f"Цена как число: {price_data}")
        return f"${price_data:.2f}"
    elif isinstance(price_data, str):
        match = re.search(r'\$?\d+(\.\d+)?', price_data)
        if match:
            logging.debug(f"Цена как строка: {match.group()}")
            return match.group()
        else:
            logging.warning("Цена не удалось извлечь из строки.")
    
    logging.warning("Цена не найдена.")
    return "Not Found"



def extract_coupon(coupon_data):
    """Извлекает значение купона из данных продукта."""
    logging.debug(f"Извлечение купона из данных: {coupon_data}")
    if isinstance(coupon_data, (int, float)):
        return f"{coupon_data}%"
    elif isinstance(coupon_data, str):
        coupon_match = re.search(r'\d+(?:\.\d{1,2})?', coupon_data)
        if coupon_match:
            return f"{float(coupon_match.group())}%"
    return 'Not Found'

def extract_bsr(product_data):
    """Извлекает Best Sellers Rank (BSR) из данных продукта."""
    logging.debug(f"Извлечение BSR из данных продукта: {product_data}")
    bsr_value = 'Not Found'
    bsr_locations = [
        'best_sellers_rank', 'bsr', 'bestsellers_rank',
        'bestseller_rank', 'sales_rank', 'rank'
    ]
    
    for location in bsr_locations:
        bsr_data = product_data.get(location)
        if bsr_data:
            logging.debug(f"Найдено BSR в '{location}': {bsr_data}")
            
            if isinstance(bsr_data, list):
                for item in bsr_data:
                    if isinstance(item, dict):
                        rank = item.get('rank') or item.get('value')
                        if rank:
                            bsr_value = str(rank).replace(',', '').split()[0]
                            break
                    elif isinstance(item, str):
                        match = re.search(r'#?([\d,]+)', item)
                        if match:
                            bsr_value = match.group(1).replace(',', '')
                            break
            elif isinstance(bsr_data, dict):
                rank = bsr_data.get('rank') or bsr_data.get('value')
                if rank:
                    bsr_value = str(rank).replace(',', '').split()[0]
            elif isinstance(bsr_data, str):
                match = re.search(r'#?([\d,]+)', bsr_data)
                if match:
                    bsr_value = match.group(1).replace(',', '')
            elif isinstance(bsr_data, (int, float)):
                bsr_value = str(int(bsr_data))
            
            if bsr_value != 'Not Found':
                break

    logging.info(f"Извлеченный BSR: {bsr_value}")
    return bsr_value

def extract_data_from_json(response_json, asin, is_variation=False):
    logging.debug("Извлечение данных из JSON.")
    try:
        product_data = response_json['results'][0]['content']
        logging.debug(f"Полные данные продукта: {json.dumps(product_data, indent=2)}")

        title = product_data.get('title', 'Not Found')
        rating = product_data.get('rating', 'Not Found')
        reviews_count = product_data.get('reviews_count') or product_data.get('review_count', 'Not Found')
        brand = product_data.get('brand', 'Not Found')
        bsr = extract_bsr(product_data)

        price = extract_price(product_data.get('price'))
        prime_price = extract_price(product_data.get('prime_offer_price'))  # Используем новый ключ
        title_price = extract_price(product_data.get('title_price'))
        
        # Извлечение List Price напрямую через price_strikethrough
        list_price = product_data.get('price_strikethrough', 'Not Found')
        logging.info(f"Извлеченный List Price: {list_price}")

        coupon = extract_coupon(product_data.get('coupon'))
        final_price = calculate_final_price(price, prime_price or price, coupon)
        discount_percent = calculate_discount_percent(price, final_price)

        product_info = {
            "ASIN": asin,
            "Title": title,
            "Price": price,
            "Prime Price": prime_price or price,  # Используем prime_price
            "Title Price": title_price,
            "List Price": list_price,
            "Coupon Discount": coupon,
            "Final Price": final_price,
            "Discount Percent": discount_percent,
            "Rating": rating,
            "Number of Reviews": reviews_count,
            "BSR": bsr,
            "Brand": brand,
            "Scrape Date": get_kyiv_time().strftime("%d.%m.%Y"),
            "URL": product_data.get('url', 'Not Found')
        }

        logging.debug(f"Prime Offer Price: {prime_price}")  # Дополнительное логирование

        logging.info(f"Извлеченные данные для ASIN {product_info['ASIN']}: {json.dumps(product_info, indent=2)}")
        return product_info

    except Exception as e:
        logging.error(f"Ошибка при извлечении данных из JSON: {str(e)}")
        return None



def extract_list_price(product_data):
    """
    Извлекает List Price из данных продукта.
    
    :param product_data: Данные о продукте из JSON.
    :return: List Price в виде строки или "Not Available".
    """
    logging.debug(f"Извлечение List Price из данных продукта: {json.dumps(product_data, indent=2)}")
    
    # Проверяем различные возможные ключи для List Price
    possible_keys = ['list_price', 'price_strikethrough', 'was_price', 'original_price', 'old_price']
    
    for key in possible_keys:
        list_price = product_data.get(key)
        if list_price:
            logging.info(f"Найден List Price под ключом '{key}': {list_price}")
            
            if isinstance(list_price, dict):
                # Если List Price - это словарь, пробуем извлечь значение
                for subkey in ['value', 'amount', 'price', 'raw']:
                    if subkey in list_price:
                        return format_price(list_price[subkey])
            elif isinstance(list_price, (int, float)):
                return format_price(list_price)
            elif isinstance(list_price, str):
                # Если это строка, пытаемся извлечь числовое значение
                match = re.search(r'\$?(\d+(?:\.\d{2})?)', list_price)
                if match:
                    return format_price(float(match.group(1)))
    
    # Если List Price не найден, выводим все ключи для анализа
    logging.warning(f"List Price не найден в данных продукта. Доступные ключи: {list(product_data.keys())}")
    return "Not Available"




def format_price(price):
    """Форматирует цену в строку с двумя знаками после запятой."""
    return f"${price:.2f}" if isinstance(price, (int, float)) else str(price)


def scrape_amazon_product(url, config, is_variation=False):
    """Скрапит данные о продукте с Amazon через Oxylabs."""
    if not url.startswith('http'):
        logging.error(f"Invalid URL: {url}")
        return None

    asin = extract_asin(url)
    if asin == 'Not Found':
        logging.error(f"ASIN not found in URL: {url}")
        return None

    payload = {
        'source': 'amazon',
        'url': url,  # Используем полный URL вместо ASIN
        'parse': True
    }

    logging.debug(f"Payload: {payload}")

    max_retries = 3
    for attempt in range(max_retries):
        try:
            api_limiter.wait()

            oxylabs_username = config.get('oxylabs_username', '').strip()
            oxylabs_password = config.get('oxylabs_password', '').strip()

            if not oxylabs_username or not oxylabs_password:
                logging.error("Oxylabs credentials are missing in the configuration")
                return None

            logging.info(f"Sending request to Oxylabs for ASIN: {asin}")

            response = requests.post(
                'https://realtime.oxylabs.io/v1/queries',
                auth=(oxylabs_username, oxylabs_password),
                json=payload,
                timeout=30
            )

            logging.debug(f"Received response: {response.status_code} - {response.text}")

            if response.status_code == 204:
                logging.error(f"No Content for ASIN {asin}")
                return None

            if response.status_code != 200:
                logging.error(f"Non-200 response from Oxylabs: {response.status_code}")
                continue

            try:
                response_json = response.json()
                logging.debug(f"Received JSON for ASIN {asin}: {json.dumps(response_json, indent=2, ensure_ascii=False)}")  # Логирование JSON-ответа
            except ValueError:
                logging.error(f"Ошибка декодирования JSON для ASIN {asin}")
                continue

            if 'error' in response_json:
                logging.error(f"Error from Oxylabs for ASIN {asin}: {response_json['error']}")
                return None

            product_info = extract_data_from_json(response_json, asin, is_variation=is_variation)
            if product_info:
                logging.info(f"Successfully scraped data for ASIN: {asin}")
                logging.debug(f"Product Info: {product_info}")
                return product_info
            else:
                logging.warning(f"Не удалось извлечь данные для ASIN {asin}")
        except requests.exceptions.RequestException as e:
            logging.error(f"Request exception for ASIN {asin}: {str(e)}")
            if attempt < max_retries - 1:
                wait_time = 5 * (attempt + 1)
                logging.info(f"Retrying in {wait_time} seconds...")
                time.sleep(wait_time)
            else:
                logging.error(f"Failed to retrieve data for ASIN {asin} after {max_retries} attempts.")
                return None

    return None


def send_telegram_message(bot, chat_id, message):
    """Отправляет сообщение в Telegram."""
    try:
        bot.send_message(chat_id, message)
        logging.info(f"Отправлено уведомление в Telegram")
    except Exception as e:
        logging.error(f"Не удалось отправить уведомление в Telegram: {str(e)}")

def check_product_notifications(product_info, min_rating, price_threshold, coupon_threshold):
    """Проверяет условия для отправки уведомлений."""
    notifications = []

    # Проверка рейтинга
    rating = product_info.get("Rating", "Not Found")
    if rating != "Not Found":
        try:
            rating_value = float(rating)
            if rating_value < min_rating:
                notifications.append(f"⚠️ Низкий рейтинг: {rating} звезд для ASIN {product_info['ASIN']}")
        except ValueError:
            notifications.append(f"⚠️ Некорректный рейтинг для ASIN {product_info['ASIN']}")

    # Проверка изменения цены
    full_price = product_info.get("Price", "Not Found")
    prime_price = product_info.get("Prime Price", "Not Found")
    if full_price != "Not Found" and prime_price != "Not Found":
        try:
            full_price_value = float(re.sub(r'[^\d.]', '', str(full_price))) if isinstance(full_price, (str, float)) else None
            prime_price_value = float(re.sub(r'[^\d.]', '', str(prime_price))) if isinstance(prime_price, (str, float)) else None
            if full_price_value and prime_price_value:
                price_diff_percent = abs(full_price_value - prime_price_value) / full_price_value * 100
                if price_diff_percent >= price_threshold:
                    notifications.append(f"💰 Значительное изменение цены для ASIN {product_info['ASIN']}: Full ${full_price_value:.2f}, Prime ${prime_price_value:.2f}")
        except ValueError:
            notifications.append(f"⚠️ Некорректная цена для ASIN {product_info['ASIN']}")

    # Проверка купона
    coupon_discount = product_info.get("Coupon Discount", "Not Found")
    if coupon_discount != "Not Found":
        try:
            coupon_value = float(re.sub(r'[^\d.]', '', str(coupon_discount).replace('%', '')))
            if coupon_value >= coupon_threshold:
                notifications.append(f"🏷️ Большой купон для ASIN {product_info['ASIN']}: {coupon_value}%")
        except ValueError:
            notifications.append(f"⚠️ Некорректная скидка купона для ASIN {product_info['ASIN']}")

    return notifications


def update_monitoring_sheet(spreadsheet, data, current_time_slot, config):
    sheet_name = "SS+Sox"  # Название листа в Google Sheets
    try:
        sheet = spreadsheet.worksheet(sheet_name)
    except gspread.exceptions.WorksheetNotFound:
        logging.error(f"Лист '{sheet_name}' не найден в таблице.")
        return

    header = [
        "Наименование", "Параметры", "Данные"
    ] + config.get('active_trade_slots', []) + ["Зона анализа"] + config.get('analysis_slots', [])

    header = [slot.strip() if isinstance(slot, str) else slot for slot in header]

    logging.debug(f"Текущий временной слот: '{current_time_slot}'")
    logging.debug(f"Заголовки таблицы: {header}")

    value_ranges = [{
        'range': f'{sheet_name}!A1',
        'values': [header]
    }]

    start_row = 3
    current_row = start_row

    # Получаем текущее время в формате "YYYY-MM-DD HH:MM:SS" в киевском часовом поясе
    current_time = get_kyiv_time().strftime("%Y-%m-%d %H:%M:%S")

    # Добавляем запись текущего времени на 2-ю строку в колонку "Данные"
    time_notation = f'{sheet_name}!C2'  # Ячейка C2 (строка 2, колонка "Данные")
    value_ranges.append({
        'range': time_notation,
        'values': [[current_time]]  # Записываем текущее время
    })


    parameters = ["BSR", "Number of Reviews", "Rating", "Price"]

    companies = [
        ("product_urls", "Merino.tech. (Мы)"),
        ("1competitor_urls", config.get('competitor_1_name', 'Competitor 1').strip()),
        ("2competitor_urls", config.get('competitor_2_name', 'Competitor 2').strip()),
    ]
    companies = [(section, name) for section, name in companies if name]

    data_to_write = []

    data_to_write.append(["Parent ASIN"])
    current_row += 1

    asin_row_mapping_parent = {}

    for param in parameters:
        data_to_write.append([param])  
        current_row += 1

        for section, company_name in companies:
            urls = config.get(section, [])
            if urls:
                for url in urls:
                    asin = extract_asin(url)
                    logging.debug(f"Обработка URL: {url}, извлеченный ASIN: {asin}, компания: {company_name}")

                    product_info = next((prod for prod in data.get(company_name, []) if prod.get('ASIN') == asin), {})

                    if not product_info:
                        asin_display = 'Not Found'
                        param_value = None
                    else:
                        asin_display = asin
                        param_value = product_info.get(param, None)
                        if isinstance(param_value, str) and param_value.lower() == "not found":
                            param_value = None

                    hyperlink_formula = f'=HYPERLINK("{url}", "{asin_display}")' if asin_display != "Not Found" else asin_display
                    data_to_write.append([company_name, hyperlink_formula, param_value])
                    asin_row_mapping_parent[(company_name, asin, param)] = current_row
                    current_row += 1

                data_to_write.append([''])
                current_row += 1

    data_to_write.append(["Variations ASIN"])
    current_row += 1

    variations_sections = [
        ("variation_urls", "Merino.tech. (Мы)"),
        ("1variation_urls", config.get('competitor_1_name', 'Competitor 1').strip()),
        ("2variation_urls", config.get('competitor_2_name', 'Competitor 2').strip()),
    ]
    variations_sections = [(section, name) for section, name in variations_sections if name]

    price_types = ["Price", "List Price", "Prime Price"]  # Убедитесь, что "Prime Price" включён

    asin_row_mapping_variations = {}
    average_row_mapping = {}

    for price_type in price_types:
        data_to_write.append([price_type])
        current_row += 1

        for section, company_name in variations_sections:
            urls = config.get(section, [])
            if urls:
                temp_asin_rows = []
                num_variations = 0

                formula_row = current_row  

                # Добавляем строку для средней цены в колонку "Данные"
                data_to_write.append([company_name, "Средняя цена", ""])  
                current_row += 1

                # Сохраняем номер строки для формулы средней цены
                average_row_number = current_row - 1
                average_row_mapping[(company_name, price_type)] = average_row_number

                for url in urls:
                    asin = extract_asin(url)
                    logging.debug(f"Обработка вариации URL: {url}, извлеченный ASIN: {asin}, компания: {company_name}")

                    product_info = next((prod for prod in data.get(company_name, []) if prod.get('ASIN') == asin), {})

                    if not product_info:
                        price_value = None
                    else:
                        price_value = product_info.get(price_type, None)
                        if isinstance(price_value, str) and price_value.lower() == "not found":
                            price_value = None

                    hyperlink_formula = f'=HYPERLINK("{url}", "{asin}")' if asin != "Not Found" else asin

                    if isinstance(price_value, (int, float)):
                        price_display = f"${price_value:.2f}"
                    elif isinstance(price_value, str):
                        price_display = price_value
                    else:
                        price_display = None

                    temp_asin_rows.append([company_name, hyperlink_formula, price_display])
                    asin_row_mapping_variations[(company_name, asin, price_type)] = current_row
                    current_row += 1
                    num_variations += 1

                data_to_write.extend(temp_asin_rows)

                # Обновляем формулу средней цены в колонке "Данные" (C-колонка)
                first_price_row = formula_row + 1
                last_price_row = formula_row + num_variations
                price_range = f"C{first_price_row}:C{last_price_row}"
                average_formula = f'=AVERAGE(FILTER({price_range}, {price_range}<>""))'

                data_to_write[formula_row - start_row][2] = average_formula

                data_to_write.append([''])
                current_row += 1

    end_row = start_row + len(data_to_write) - 1
    range_notation = f'{sheet_name}!A{start_row}:C{end_row}'

    value_ranges.append({
        'range': range_notation,
        'values': data_to_write
    })

    if current_time_slot and current_time_slot in header:
        slot_column = header.index(current_time_slot) + 1  
        logging.info(f"Данные будут записаны в колонку '{current_time_slot}' (столбец {slot_column})")
    else:
        logging.info("Текущее время не совпадает с временными слотами, данные будут записаны в колонку 'Данные'")
        slot_column = header.index("Данные") + 1

    column_letter = get_column_letter(slot_column)

    slot_updates = []

    def find_product_info(all_data, company_name, asin):
        products = all_data.get(company_name, [])
        for prod in products:
            if prod.get('ASIN') == asin:
                return prod
        return None

    for (company_name, asin, param), row_number in asin_row_mapping_parent.items():
        product_info = find_product_info(data, company_name, asin)
        if product_info:
            value = product_info.get(param, None)
            if isinstance(value, str) and value.lower() == "not found":
                value = None  
            cell_notation = f'{sheet_name}!{column_letter}{row_number}'
            slot_updates.append({
                'range': cell_notation,
                'values': [[value]]
            })

    for (company_name, asin, price_type), row_number in asin_row_mapping_variations.items():
        product_info = find_product_info(data, company_name, asin)
        if product_info:
            price_value = product_info.get(price_type, None)
            if isinstance(price_value, str) and price_value.lower() == "not found":
                price_value = None  
            elif isinstance(price_value, (int, float)):
                price_value = f"${price_value:.2f}"
            cell_notation = f'{sheet_name}!{column_letter}{row_number}'
            slot_updates.append({
                'range': cell_notation,
                'values': [[price_value]]
            })

    # Обновление формулы средней цены в соответствующем временном слоте
    for (company_name, price_type), row_number in average_row_mapping.items():
        variation_rows = [row_num for (comp, asin, ptype), row_num in asin_row_mapping_variations.items()
                          if comp == company_name and ptype == price_type]

        if variation_rows:
            first_row = min(variation_rows)
            last_row = max(variation_rows)
            slot_column_letter = column_letter  

            price_range = f'{slot_column_letter}{first_row}:{slot_column_letter}{last_row}'

            average_formula = f'=AVERAGE(FILTER({price_range}, {price_range}<>""))'

            cell_notation = f'{sheet_name}!{slot_column_letter}{row_number}'

            slot_updates.append({
                'range': cell_notation,
                'values': [[average_formula]]
            })

    value_ranges.extend(slot_updates)

    try:
        data_body = {
            'value_input_option': 'USER_ENTERED',
            'data': value_ranges
        }
        spreadsheet.values_batch_update(data_body)
        logging.info("Данные успешно обновлены в Google Sheets.")
    except APIError as e:
        logging.error(f"Ошибка API при обновлении Google Sheets: {str(e)}")
        retry_delay = 60  
        logging.info(f"Попытка повторного обновления через {retry_delay} секунд...")
        time.sleep(retry_delay)
        try:
            spreadsheet.values_batch_update(data_body)
            logging.info("Данные успешно обновлены в Google Sheets после повторной попытки.")
        except Exception as e2:
            logging.error(f"Не удалось обновить Google Sheets после повторной попытки: {str(e2)}")
    except Exception as e:
        logging.error(f"Ошибка при обновлении Google Sheets: {str(e)}")







def extract_reviews_count(product_data):
    """Извлекает количество отзывов из данных продукта."""
    reviews_count = 'Not Found'
    possible_keys = ['reviews_count', 'rating_count', 'ratings_total', 'review_count']
    
    for key in possible_keys:
        if key in product_data:
            reviews_count = product_data[key]
            break
    
    if reviews_count != 'Not Found':
        if isinstance(reviews_count, str):
            reviews_count = re.sub(r'[^\d]', '', reviews_count)
        
        try:
            reviews_count = int(reviews_count)
        except ValueError:
            reviews_count = 'Not Found'
   
    logging.info(f"Извлеченное количество отзывов: {reviews_count}")
    return reviews_count

def create_xlsx_report(data, current_time_str):
    """Создание XLSX отчета с данными о продуктах."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Product Report"

        # Стили
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        bold_blue_font = Font(bold=True, color="0000FF")  # Жирный и синий цвет для кликабельных ASIN

        # Заголовок
        ws['A1'] = f"Product Report - {current_time_str}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:N1')  # Расширил диапазон для новых колонок

        # Заголовки столбцов
        headers = [
            "Company", "ASIN", "Title", "Full Price", "Prime Price", 
            "Avg List Price", "Avg Title Price", "Avg Prime Price",
            "Rating", "Number of Reviews", "Coupon Discount", 
            "Final Price", "Discount Percent", "Variations Count"
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Данные
        row = 3
        for company, products in data.items():
            for product in products:
                ws.cell(row=row, column=1, value=company)
                asin_cell = ws.cell(row=row, column=2, value=product.get('ASIN', 'Не найдено'))
                asin = product.get('ASIN', 'Не найдено')
                if asin != 'Не найдено':
                    asin_cell.hyperlink = f"https://www.amazon.com/dp/{asin}"
                    asin_cell.font = bold_blue_font  # Сделать кликабельным ASIN жирным и синим
                ws.cell(row=row, column=3, value=product.get('Title', 'Не найдено'))
                ws.cell(row=row, column=4, value=product.get('Price', 'Не найдено'))
                ws.cell(row=row, column=5, value=product.get('Prime Price', 'Не найдено'))
                ws.cell(row=row, column=6, value=product.get('List Price', 'Not Found'))
                ws.cell(row=row, column=7, value=product.get('Sale Price', 'Not Found'))
                ws.cell(row=row, column=8, value=product.get('Prime Price', 'Not Found'))
                ws.cell(row=row, column=9, value=product.get('Rating', 'Не найдено'))
                ws.cell(row=row, column=10, value=product.get('Number of Reviews', 'Не найдено'))
                ws.cell(row=row, column=11, value=product.get('Coupon Discount', 'Not Found'))
                ws.cell(row=row, column=12, value=product.get('Final Price', 'Not Found'))
                ws.cell(row=row, column=13, value=product.get('Discount Percent', 'Not Found'))
                ws.cell(row=row, column=14, value=product.get('Variations Count', 'Not Found'))
                row += 1

        # Автоматическая регулировка ширины столбцов
        for col in ws.columns:
            max_length = 0
            column = col[0].column  # Получаем индекс столбца
            column_letter = get_column_letter(column)
            for cell in col:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Сохранение файла
        filename = f"product_report_{current_time_str.replace(':', '_')}.xlsx"
        wb.save(filename)
        return filename
    except Exception as e:
        logging.error(f"Ошибка при создании XLSX отчета: {str(e)}")
        return None

def send_telegram_notification(config, current_time_str, data):
    """Отправка уведомления и отчета в Telegram."""
    bot = telebot.TeleBot(config.get('telegram_bot_token', ''))
    chat_id = config.get('telegram_chat_id', '')
    if not chat_id:
        logging.error("Telegram chat_id не установлен в конфигурации.")
        return

    # Формирование сообщения
    message = f"Отчет за {current_time_str}\n\n"
    for company, products in data.items():
        message += f"{company}:\n"
        for product in products:
            message += f" ASIN: {product['ASIN']}\n"
            message += f" Цена: {product.get('Price', 'Не найдено')}\n"
            message += f" Рейтинг: {product.get('Rating', 'Не найдено')}\n\n"

    # Отправка сообщения
    send_telegram_message(bot, chat_id, message)

    # Создание и отправка XLSX файла
    xlsx_filename = create_xlsx_report(data, current_time_str)
    if xlsx_filename:
        try:
            with open(xlsx_filename, 'rb') as report_file:
                bot.send_document(chat_id, report_file)
            logging.info("Отчет успешно отправлен в Telegram.")
        except Exception as e:
            logging.error(f"Не удалось отправить отчет в Telegram: {str(e)}")
        finally:
            if os.path.exists(xlsx_filename):
                os.remove(xlsx_filename)

def round_time_to_nearest_slot(current_time_str, active_trade_slots, analysis_slots):
    """Округляет текущее время до ближайшего временного слота."""
    all_slots = active_trade_slots + analysis_slots
    if not all_slots:
        logging.error("Нет доступных временных слотов для сравнения.")
        return None

    current_time_minutes = int(current_time_str.split(":")[0]) * 60 + int(current_time_str.split(":")[1])
    
    nearest_slot = min(all_slots, key=lambda slot: abs(current_time_minutes - (int(slot.split(":")[0]) * 60 + int(slot.split(":")[1]))))
    return nearest_slot

def update_google_sheets(current_results, spreadsheet_id, config, current_time_str, credentials_file):
    """Обновление Google Sheets данными из current_results."""
    try:
        client = authorize_google_sheets(credentials_file)
        spreadsheet = client.open_by_key(spreadsheet_id)
        
        # Получаем текущее время в формате HH:MM
        current_time_formatted = current_time_str.split(' ')[1][:5]
        
        all_slots = config.get('active_trade_slots', []) + config.get('analysis_slots', [])
        
        if current_time_formatted in all_slots:
            current_time_slot = current_time_formatted
        else:
            current_time_slot = None  # Текущее время не совпадает с временными слотами

        update_monitoring_sheet(spreadsheet, current_results, current_time_slot, config)

    except APIError as e:
        logging.error(f"Ошибка API при обновлении Google Sheets: {str(e)}")

def gather_product_data(config, competitor_urls, competitor_variation_urls):
    """Функция для сбора данных по продуктам. Возвращает текущие результаты."""
    current_results = {
        "Merino.tech. (Мы)": [],
        "Merino Protect": [],
        "METARINO": [],
    }

    # Обработка основных продуктов (Parent ASIN)
    logging.info("Начинаем сбор данных по основным продуктам Merino.tech (Parent ASIN).")
    for url in config.get('product_urls', []):
        logging.debug(f"Обработка Parent ASIN по URL: {url}")
        try:
            product_info = scrape_amazon_product(url, config)
            if product_info:
                logging.info(f"Успешно собраны данные для основного продукта (Parent ASIN): {url}")
                current_results["Merino.tech. (Мы)"].append(product_info)
            else:
                logging.warning(f"Не удалось получить данные для основного продукта (Parent ASIN): {url}")
        except Exception as e:
            logging.error(f"Ошибка при сборе данных для Parent ASIN по продукту {url}: {str(e)}")

    # Обработка вариаций продуктов (Variation ASIN)
    logging.info("Начинаем сбор данных по вариациям продуктов Merino.tech (Variation ASIN).")
    for var_url in config.get('variation_urls', []):
        logging.debug(f"Обработка Variation ASIN по URL: {var_url}")
        try:
            variation_info = scrape_amazon_product(var_url, config, is_variation=True)
            if variation_info:
                logging.info(f"Успешно собраны данные для вариации продукта (Variation ASIN): {var_url}")
                current_results["Merino.tech. (Мы)"].append(variation_info)
            else:
                logging.warning(f"Не удалось получить данные для вариации продукта: {var_url}")
        except Exception as e:
            logging.error(f"Ошибка при сборе данных для Variation ASIN по продукту {var_url}: {str(e)}")

    # Обработка конкурентов
    logging.info("Начинаем сбор данных по конкурентам.")
    for competitor_name, urls in competitor_urls.items():
        for url in urls:
            logging.debug(f"Обработка продукта конкурента по URL: {url}")
            try:
                product_info = scrape_amazon_product(url, config)
                if product_info:
                    logging.info(f"Успешно собраны данные для конкурента {competitor_name}: {url}")
                    current_results.setdefault(competitor_name, []).append(product_info)
                else:
                    logging.warning(f"Не удалось получить данные для конкурента {competitor_name}: {url}")
            except Exception as e:
                logging.error(f"Ошибка при сборе данных для конкурента {competitor_name} ({url}): {str(e)}")

    # Обработка вариаций конкурентов
    logging.info("Начинаем сбор данных по вариациям конкурентов.")
    for competitor_name, var_urls in competitor_variation_urls.items():
        for var_url in var_urls:
            logging.debug(f"Обработка вариации продукта конкурента по URL: {var_url}")
            try:
                variation_info = scrape_amazon_product(var_url, config, is_variation=True)
                if variation_info:
                    logging.info(f"Успешно собраны данные для вариации конкурента {competitor_name}: {var_url}")
                    current_results.setdefault(competitor_name, []).append(variation_info)
                else:
                    logging.warning(f"Не удалось получить данные для вариации конкурента {competitor_name}: {var_url}")
            except Exception as e:
                logging.error(f"Ошибка при сборе данных для вариации конкурента {competitor_name} ({var_url}): {str(e)}")

    return current_results


def find_credentials_file():
    """Пытается найти файл учетных данных по нескольким возможным путям."""
    possible_paths = [
        os.path.join(os.path.expanduser('~'), 'Downloads', 'maximumstores53-24d4ef8c1298.json'),
        os.path.join(os.getcwd(), 'maximumstores53-24d4ef8c1298.json'),
        'maximumstores53-24d4ef8c1298.json'
    ]
    
    for path in possible_paths:
        if os.path.exists(path):
            logging.info(f"Найден файл учетных данных: {path}")
            return path

    logging.error("Файл учетных данных не найден ни в одном из возможных путей.")
    return None

def main():
    """Основная функция скрипта."""
    # Поиск файла учетных данных Google Sheets
    credentials_file = find_credentials_file()
    if not credentials_file:
        logging.critical("Не удалось найти файл учетных данных.")
        return

    # ID таблицы Google Sheets
    spreadsheet_id = '1ibuYnN9WeRZdHUqoiU2jFLez59fm5Gfgzeyvq7M4EaI'

    # Авторизация и загрузка конфигурации
    try:
        client = authorize_google_sheets(credentials_file)
        config = load_config_from_sheets(client, spreadsheet_id)
    except Exception as e:
        logging.critical(f"Не удалось авторизоваться или загрузить конфигурацию: {e}")
        return

    # Логирование для проверки конфигурации
    logging.info(f"Product URLs after loading config: {config.get('product_urls', [])}")
    logging.info(f"Variation URLs after loading config: {config.get('variation_urls', [])}")
    logging.info(f"Competitor 1 URLs after loading config: {config.get('1competitor_urls', [])}")
    logging.info(f"Competitor 2 URLs after loading config: {config.get('2competitor_urls', [])}")
    logging.info(f"Competitor 3 URLs after loading config: {config.get('3competitor_urls', [])}")
    logging.info(f"Variation URLs for competitor 1 (Merino Protect): {config.get('1variation_urls', [])}")
    logging.info(f"Variation URLs for competitor 2 (METARINO): {config.get('2variation_urls', [])}")

    # Инициализация Telegram бота
    telegram_bot_token = config.get('telegram_bot_token', '')
    telegram_chat_id = config.get('telegram_chat_id', '')
    if not telegram_bot_token or not telegram_chat_id:
        logging.critical("Telegram bot token или chat_id не установлены в конфигурации.")
        return
    bot = telebot.TeleBot(telegram_bot_token)

    # Получение URL-адресов продуктов и вариаций
    PRODUCT_URLS = config.get('product_urls', [])
    VARIATION_URLS = config.get('variation_urls', [])

    # Определение конкурентных продуктов и вариаций на основе конфигурации
    COMPETITOR_URLS = {
        "Merino Protect": config.get('1competitor_urls', []),
        "METARINO": config.get('2competitor_urls', []),
    }

    COMPETITOR_VARIATION_URLS = {
        "Merino Protect": config.get('1variation_urls', []),
        "METARINO": config.get('2variation_urls', []),
    }

    # Получение других настроек из конфигурации
    BATCH_SIZE = config.get('batch_size', 100)
    MIN_ACCEPTABLE_RATING = config.get('min_acceptable_rating', 4.0)
    PRICE_CHANGE_THRESHOLD = config.get('price_change_threshold', 5.0)
    COUPON_THRESHOLD = config.get('coupon_threshold', 10.0)

    # Получение времени обновления из конфигурации
    update_hour = config.get('update_time_hour', 0)
    update_minute = config.get('update_time_minute', 0)
    timezone_str = config.get('timezone', 'Europe/Kiev')

    # Получение временных слотов из конфигурации
    active_trade_slots = config.get('active_trade_slots', ["14:00", "16:00", "18:00", "20:00", "22:00", "00:00", "02:00", "04:00", "06:00"])
    analysis_slots = config.get('analysis_slots', ["08:00", "10:00", "12:00"])

    # Добавляем специальное время обновления
    update_time_slot = f"{int(update_hour):02d}:{int(update_minute):02d}"
    all_slots = active_trade_slots + analysis_slots + [update_time_slot]

    # Удаляем дубликаты и сортируем
    all_slots = sorted(list(set(all_slots)))

    logging.info(f"Все временные слоты: {all_slots}")

    # #### ТЕСТОВАЯ ВЕРСИЯ ЦИКЛА: Немедленный запуск процесса ####
    try:
        # Получаем текущее время
        current_time = get_kyiv_time(timezone_str)
        logging.info(f"Текущее время: {current_time.strftime('%Y-%m-%d %H:%M:%S')}")

        # Устанавливаем текущий слот как ближайший слот (или используем текущее время)
        nearest_slot = get_next_slot(current_time, all_slots, timezone_str)
        logging.info(f"Тестовый запуск процесса для временного слота: {nearest_slot.strftime('%H:%M')}")

        # Выполнение сбора данных
        current_results = gather_product_data(config, COMPETITOR_URLS, COMPETITOR_VARIATION_URLS)

        # Обновление Google Sheets
        update_google_sheets(
            current_results,
            spreadsheet_id,
            config,
            nearest_slot.strftime('%Y-%m-%d %H:%M:%S'),
            credentials_file
        )

        # Отправка уведомлений в Telegram
        if current_results:
            send_telegram_notification(
                config,
                nearest_slot.strftime('%Y-%m-%d %H:%M:%S'),
                current_results
            )

        logging.info(f"Тестовый процесс для слота {nearest_slot.strftime('%H:%M')} успешно завершен.")

    except Exception as e:
        logging.error(f"Ошибка при тестовом запуске процесса: {e}")

    # #### Оригинальный ЦИКЛ: Регулярные Запуски ####
    while True:
        try:
            current_time = get_kyiv_time(timezone_str)
            logging.info(f"Текущее время: {current_time.strftime('%Y-%m-%d %H:%M:%S')}")

            # Определение ближайшего следующего слота
            next_slot_time = get_next_slot(current_time, all_slots, timezone_str)
            logging.info(f"Следующий слот: {next_slot_time.strftime('%Y-%m-%d %H:%M')}")

            # Вычисление времени до следующего слота
            time_to_wait = (next_slot_time - current_time).total_seconds()

            if time_to_wait > 0:
                logging.info(f"Ждем {time_to_wait / 60:.2f} минут до следующего слота.")
                time.sleep(time_to_wait)

            # После пробуждения запустить задачу
            logging.info(f"Запуск процесса для временного слота: {next_slot_time.strftime('%H:%M')}")

            # Выполнение сбора данных
            current_results = gather_product_data(config, COMPETITOR_URLS, COMPETITOR_VARIATION_URLS)

            # Обновление Google Sheets
            update_google_sheets(
                current_results,
                spreadsheet_id,
                config,
                next_slot_time.strftime('%Y-%m-%d %H:%M:%S'),
                credentials_file
            )

            # Отправка уведомлений в Telegram
            if current_results:
                send_telegram_notification(
                    config,
                    next_slot_time.strftime('%Y-%m-%d %H:%M:%S'),
                    current_results
                )

            logging.info(f"Процесс для слота {next_slot_time.strftime('%H:%M')} успешно завершен.")

        except Exception as e:
            logging.error(f"Ошибка в основном цикле: {e}")
            time.sleep(60)  # Ждем минуту перед повторной попыткой
    # #### КОНЕЦ ОРИГИНАЛЬНОГО ЦИКЛА ####


if __name__ == '__main__':
    main()
