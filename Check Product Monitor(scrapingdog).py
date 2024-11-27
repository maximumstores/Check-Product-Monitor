import time
import logging
from datetime import datetime, timedelta
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import requests
import re
from urllib.parse import urlparse, parse_qs
import os
from threading import Lock
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import telebot
import random
import pytz
import json
from gspread.exceptions import APIError
from bs4 import BeautifulSoup  # Для парсинга HTML
from google.oauth2.service_account import Credentials
from gspread_formatting import CellFormat, format_cell_range, Color, TextFormat
import sys
from logging.handlers import RotatingFileHandler

# Настройка базового конфигуратора логирования
logger = logging.getLogger()
logger.setLevel(logging.DEBUG)  # Устанавливаем уровень логирования на DEBUG для подробных логов

# Создаем форматтер
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

# Создаем обработчик для записи логов в файл с ротацией
file_handler = RotatingFileHandler('scraper.log', maxBytes=5*1024*1024, backupCount=5, encoding='utf-8')
file_handler.setLevel(logging.DEBUG)  # Логируем все уровни
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)

# Создаем обработчик для вывода логов в терминал
console_handler = logging.StreamHandler(sys.stdout)
console_handler.setLevel(logging.INFO)  # В терминал выводим INFO и выше
console_handler.setFormatter(formatter)
logger.addHandler(console_handler)

class APIRateLimiter:
    """Класс для ограничения количества запросов к API."""
    def __init__(self, max_requests, period):
        self.max_requests = max_requests
        self.period = period
        self.requests = []
        self.lock = Lock()

    def wait(self):
        """Ожидание перед следующим запросом, если достигнут лимит."""
        with self.lock:
            now = time.time()
            # Удаляем устаревшие запросы
            self.requests = [r for r in self.requests if r > now - self.period]
            while len(self.requests) >= self.max_requests:
                next_request_time = self.requests[0] + self.period
                sleep_time = max(next_request_time - now, 0)
                logging.debug(f"Достигнут лимит запросов. Спим {sleep_time:.2f} секунд.")
                time.sleep(sleep_time)
                now = time.time()
                self.requests = [r for r in self.requests if r > now - self.period]
            self.requests.append(now)

# Настройка лимитера на 1 запрос в секунду
api_limiter = APIRateLimiter(max_requests=1, period=1)  # 1 запрос в секунду

def clean_urls(raw_value):
    """
    Очищает строку URL-адресов, корректно обрабатывает любые разделители,
    включая переносы строк, запятые, пробелы, а также разделители с запятыми и переносами строк.
    """
    if isinstance(raw_value, str):
        # Разделяем по переносам строк, запятым и пробелам
        urls = [url.strip() for url in re.split(r'[\n\r,]+', raw_value) if url.strip()]
    else:
        urls = raw_value if isinstance(raw_value, list) else []
    return urls

def authorize_google_sheets(credentials_file):
    """
    Авторизуется в Google Sheets и возвращает клиентский объект.
    
    :param credentials_file: Путь к файлу учетных данных JSON
    :return: gspread.Client объект
    """
    scopes = [
        'https://www.googleapis.com/auth/spreadsheets',
        'https://www.googleapis.com/auth/drive'
    ]
    
    credentials = Credentials.from_service_account_file(credentials_file, scopes=scopes)
    client = gspread.authorize(credentials)
    logging.info("Успешно авторизовались в Google Sheets")
    return client

def load_config_from_sheets(client, spreadsheet_id, config_sheet_name=None):
    """
    Загрузка конфигурации из Google Sheets.
    Если config_sheet_name не указан, загружается основной конфиг из листа 'Config'.
    """
    if not config_sheet_name:
        config_sheet_name = 'Config'

    try:
        config_sheet = client.open_by_key(spreadsheet_id).worksheet(config_sheet_name)
    except gspread.exceptions.WorksheetNotFound:
        logging.error(f"Лист '{config_sheet_name}' не найден в таблице.")
        raise

    config = {}
    all_records = config_sheet.get_all_records()

    for record in all_records:
        key = str(record.get('Key', '')).strip()
        value = str(record.get('Value', '')).strip()

        if not key:
            logging.warning("Пропущена запись с пустым ключом.")
            continue

        # Обработка URL-полей
        url_keys = [
            'product_urls', '1competitor_urls', '2competitor_urls',
            '3competitor_urls', '4competitor_urls', '5competitor_urls',
            'variation_urls', '1variation_urls', '2variation_urls',
            '3variation_urls', '4variation_urls', '5variation_urls'
        ]

        if key in url_keys:
            config[key] = clean_urls(value)
            logging.debug(f"Загружены URL для '{key}': {config[key]}")
        elif key in ['update_time_hour', 'update_time_minute', 'batch_size']:
            try:
                config[key] = int(value)
                logging.debug(f"Загружено целое число для '{key}': {config[key]}")
            except ValueError:
                logging.error(f"Некорректное целое число для '{key}': {value}. Установлено значение по умолчанию 0.")
                config[key] = 0
        elif key in ['min_acceptable_rating', 'price_change_threshold', 'coupon_threshold']:
            try:
                config[key] = float(value)
                logging.debug(f"Загружено число с плавающей точкой для '{key}': {config[key]}")
            except ValueError:
                logging.error(f"Некорректное число с плавающей точкой для '{key}': {value}. Установлено значение по умолчанию 0.0.")
                config[key] = 0.0
        elif key in ['active_trade_slots', 'analysis_slots']:
            if value:
                # Разделяем слоты по запятым и пробелам
                config[key] = [slot.strip() for slot in re.split(r'[,\s]+', value) if slot.strip()]
                logging.debug(f"Загружены временные слоты для '{key}': {config[key]}")
            else:
                config[key] = []
        elif key in ['ScrapingDogAPIKey', 'telegram_bot_token', 'telegram_chat_id', 'company_name']:
            config[key] = value
            logging.debug(f"Загружено значение для '{key}': {config[key]}")
        else:
            config[key] = value
            logging.debug(f"Загружено значение для '{key}': {config[key]}")

    # Обновление имен конкурентов
    competitor_names = {}
    for i in range(1, 6):
        competitor_name_key = f'competitor_{i}_name'
        competitor_name = config.get(competitor_name_key, '').strip()
        if competitor_name:
            competitor_names[str(i)] = competitor_name
            logging.debug(f"Добавлено имя конкурента {i}: {competitor_name}")

    config['competitor_names'] = competitor_names

    # Извлечение имени нашей компании
    company_name = config.get('company_name', '').strip()
    if not company_name:
        # Устанавливаем значение по умолчанию, если имя компании не задано
        company_name = 'Merino.tech. (Мы)'
    config['company_name'] = company_name
    logging.debug(f"Установлено имя компании: {company_name}")

    logging.info(f"Загруженная конфигурация из '{config_sheet_name}': {json.dumps(config, indent=2, ensure_ascii=False)}")
    return config

def extract_asin(url):
    """Извлекает ASIN из различных форматов URL Amazon."""
    # Парсим URL
    parsed_url = urlparse(url)
    path = parsed_url.path

    # Регулярное выражение для извлечения ASIN из пути URL
    patterns = [
        r'/dp/([A-Z0-9]{10})',
        r'/gp/product/([A-Z0-9]{10})',
        r'/product/([A-Z0-9]{10})',
        r'/ASIN/([A-Z0-9]{10})',
        r'/gp/aw/d/([A-Z0-9]{10})',
        r'/gp/offer-listing/([A-Z0-9]{10})',
    ]
    for pattern in patterns:
        match = re.search(pattern, path)
        if match:
            asin = match.group(1)
            logging.debug(f"Extracted ASIN {asin} from URL path: {url}")
            return asin
    # Если не удалось найти ASIN в пути, попробуем извлечь из параметров запроса
    query_params = parse_qs(parsed_url.query)
    if 'asin' in query_params:
        asin = query_params['asin'][0]
        logging.debug(f"Extracted ASIN {asin} from query parameters in URL: {url}")
        return asin
    # Если всё ещё не удалось найти ASIN, попробуем найти его в URL целиком
    match = re.search(r'([A-Z0-9]{10})', url)
    if match:
        asin = match.group(1)
        logging.debug(f"Extracted ASIN {asin} from entire URL: {url}")
        return asin
    logging.warning(f"Could not extract ASIN from URL: {url}")
    return 'Not Found'

def is_valid_amazon_url(url):
    """Проверяет, является ли URL корректным Amazon продуктом."""
    parsed_url = urlparse(url)
    return parsed_url.netloc in ['www.amazon.com', 'amazon.com'] and '/dp/' in parsed_url.path

def write_hyperlinks(sheet, urls, start_row, column):
    """
    Записывает гиперссылки в Google Sheets, каждая ссылка в отдельной строке.
    
    :param sheet: Лист Google Sheets
    :param urls: Список URL-адресов для записи
    :param start_row: Начальная строка для записи
    :param column: Колонка для записи (например, 'A', 'B')
    """
    for i, url in enumerate(urls):
        asin = extract_asin(url)  # Извлекаем ASIN для отображения
        if asin != 'Not Found':
            # Создаем гиперссылку для каждой ссылки отдельно
            hyperlink_formula = f'=HYPERLINK("{url}", "{asin}")'
            cell = f'{column}{start_row + i}'  # Каждая ссылка в новой строке
            sheet.update_acell(cell, hyperlink_formula)
        else:
            logging.warning(f"ASIN not found for URL: {url}")
def calculate_final_price(full_price, prime_price, coupon_discount, currency_symbol='$'):
    """ 
    Вычисляет итоговую цену с учётом скидок и купонов. 
    Возвращает строку с форматом цены.
    """
    try:
        logging.debug(f"Calculating final price with currency symbol: {currency_symbol}")

        def price_to_float(price):
            if isinstance(price, (int, float)):
                return float(price)
            if not price or price in ["Not Found", "#N/A", "#DIV/0!"]:
                return 0.0  # Возвращаем 0.0, если цена не найдена
            price = price.replace(',', '.')
            return float(re.sub(r'[^\d.]', '', price))

        full_price_value = price_to_float(full_price)
        prime_price_value = price_to_float(prime_price)

        coupon_discount_value = float(re.sub(r'[^\d.]', '', str(coupon_discount).replace('%', ''))) if coupon_discount and coupon_discount != "Not Found" else 0.0

        # Используем prime_price_value, если доступно, иначе full_price_value
        base_price = prime_price_value if prime_price_value > 0 else full_price_value

        if base_price == 0:
            logging.warning("Base price is not available for final price calculation.")
            return f"{currency_symbol}0.00"  # Возвращаем 0.0, если ни одна цена не доступна

        # Вычисляем итоговую цену с учётом купона
        discount_amount = base_price * (coupon_discount_value / 100)
        final_price_value = base_price - discount_amount
        logging.debug(f"Final price value: {final_price_value} with currency symbol: {currency_symbol}")

        # Возвращаем итоговую цену в формате строки с правильным символом валюты
        return f"{currency_symbol}{final_price_value:.2f}"
    except ValueError as e:
        logging.error(f"Ошибка при расчете цены: {str(e)}")
        return f"{currency_symbol}0.00"  # Возвращаем 0.0 в случае ошибки


def calculate_discount_percent(full_price, final_price):
    """Вычисляет процент скидки."""
    try:
        if full_price == "Not Found" or final_price == "Not Found":
            return "Не применимо"
        
        full_price_value = float(re.sub(r'[^\d.]', '', str(full_price)))
        final_price_value = float(re.sub(r'[^\d.]', '', str(final_price)))
        
        if full_price_value == 0:
            return "N/A"
       
        discount_percent_value = (full_price_value - final_price_value) / full_price_value * 100
        return f"{discount_percent_value:.2f}%"
    except ValueError:
        logging.error(f"Ошибка при расчете процента скидки с Full Price: {full_price} и Final Price: {final_price}")
        return "Не применимо"

def extract_price(price_data, currency_code='EUR'):
    """
    Извлекает и форматирует цену из данных продукта.

    :param price_data: Данные о цене из JSON-ответа или HTML
    :param currency_code: Код валюты, например 'EUR', 'USD'.
    :return: Строка с форматом цены или "Not Found".
    """
    logging.debug(f"Извлечение цены из данных: {price_data} с валютой {currency_code}")

    # Получаем символ валюты, если он известен, иначе используем код
    currency_symbol = CURRENCY_SYMBOLS.get(currency_code, currency_code)

    if not price_data or price_data == "Not Found":
        logging.warning("Цена не найдена в данных продукта.")
        return "Not Found"

    def extract_price_from_string(price_str, currency_symbol):
        """
        Извлекает цену из строки, содержащей символ валюты.
        Предполагается, что фактическая цена находится перед или рядом с символом валюты.
        """
        # Регулярное выражение для поиска числа с последующим символом валюты
        pattern_before = rf'(\d{{1,3}}(?:[.,]\d{{2}})?)\s?{re.escape(currency_symbol)}'
        pattern_after = rf'{re.escape(currency_symbol)}\s?(\d{{1,3}}(?:[.,]\d{{2}})?)'

        # Ищем совпадения с символом валюты после числа
        matches_before = re.findall(pattern_before, price_str)
        if matches_before:
            actual_price = matches_before[-1].replace(',', '.')
            try:
                price_float = float(actual_price)
                return f"{price_float:.2f} {currency_symbol}"
            except ValueError:
                logging.error(f"Ошибка преобразования найденной цены в число: {actual_price}")

        # Ищем совпадения с символом валюты перед числом
        matches_after = re.findall(pattern_after, price_str)
        if matches_after:
            actual_price = matches_after[-1].replace(',', '.')
            try:
                price_float = float(actual_price)
                return f"{price_float:.2f} {currency_symbol}"
            except ValueError:
                logging.error(f"Ошибка преобразования найденной цены в число: {actual_price}")

        logging.warning(f"Не удалось извлечь цену из строки: {price_str}")
        return "Not Found"

    # Обработка случая, если price_data - это число
    if isinstance(price_data, (int, float)):
        logging.debug(f"Цена как число: {price_data}")
        return f"{price_data:.2f} {currency_symbol}"

    # Обработка случая, если price_data - это строка
    elif isinstance(price_data, str):
        price = extract_price_from_string(price_data, currency_symbol)
        if price != "Not Found":
            return price

    logging.warning("Цена не найдена.")
    return "Not Found"

CURRENCY_SYMBOLS = {
    'EUR': '€',
    'USD': '$',
    'GBP': '£',
    'JPY': '¥',
    'CAD': 'C$',
    'AUD': 'A$',
    'CHF': 'CHF',
    'CNY': '¥',
    'RUB': '₽',
    'INR': '₹',
    'BRL': 'R$',
    'AED': 'د.إ',
    'SEK': 'kr',    # Для шведских крон
    'SGD': 'S$',    # Для сингапурских долларов
    # Добавьте другие валюты по необходимости
}

def determine_currency(url):
    parsed_url = urlparse(url)
    domain = parsed_url.netloc.lower()

    amazon_currency_mapping = {
        'amazon.com': 'USD',       # США
        'amazon.co.uk': 'GBP',     # Великобритания
        'amazon.de': 'EUR',        # Германия
        'amazon.fr': 'EUR',        # Франция
        'amazon.it': 'EUR',        # Италия
        'amazon.es': 'EUR',        # Испания
        'amazon.ca': 'CAD',        # Канада
        'amazon.co.jp': 'JPY',     # Япония
        'amazon.com.au': 'AUD',    # Австралия
        'amazon.nl': 'EUR',        # Нидерланды
        'amazon.se': 'SEK',        # Швеция
        'amazon.sg': 'SGD',        # Сингапур
        'amazon.in': 'INR',        # Индия
        'amazon.com.br': 'BRL',    # Бразилия
        'amazon.ae': 'AED',        # ОАЭ
        # Добавьте другие домены и валюты по необходимости
    }

    # Поиск соответствия домену
    for amazon_domain, currency in amazon_currency_mapping.items():
        if domain.endswith(amazon_domain):
            logging.debug(f"Домен '{domain}' соответствует валюте '{currency}'.")
            return currency

    # Если домен не найден в сопоставлении, выводим предупреждение и возвращаем символ по умолчанию
    logging.warning(f"Неизвестный домен Amazon '{domain}'. Используется символ валюты по умолчанию 'USD'.")
    return 'USD'  # Значение по умолчанию

def extract_coupon(coupon_data):
    """Извлекает значение купона из данных продукта."""
    logging.debug(f"Извлечение купона из данных: {coupon_data}")
    if isinstance(coupon_data, (int, float)):
        return f"{coupon_data}%"
    elif isinstance(coupon_data, str):
        # Пример: "43 Prozent Einsparungen"
        coupon_match = re.search(r'(\d+(?:\.\d+)?)\s*Prozent', coupon_data)
        if coupon_match:
            return f"{float(coupon_match.group(1))}%"
    return 'Not Found'

def extract_rating(rating_data):
    """Извлекает рейтинг продукта из данных."""
    logging.debug(f"Извлечение рейтинга из данных: {rating_data}")
    if isinstance(rating_data, (int, float)):
        return str(rating_data)
    elif isinstance(rating_data, str):
        # Заменяем запятую на точку для корректного преобразования
        rating_data = rating_data.replace(',', '.')
        rating_match = re.search(r'(\d+(?:\.\d+)?)', rating_data)
        if rating_match:
            return rating_match.group(1)
    return 'Not Found'

def get_kyiv_time(timezone_str='Europe/Kiev'):
    """Возвращает текущее время в часовом поясе Киева."""
    timezone = pytz.timezone(timezone_str)
    return datetime.now(timezone)

def get_next_slot(current_time, slots, timezone_str='Europe/Kiev'):
    """
    Возвращает ближайший следующий слот из списка slots.

    :param current_time: Текущее время (datetime объект)
    :param slots: Список временных слотов в формате "HH:MM"
    :param timezone_str: Часовой пояс
    :return: datetime объект ближайшего следующего слота
    """
    timezone = pytz.timezone(timezone_str)
    today_slots = sorted(slots)
    for slot in today_slots:
        slot_time = datetime.strptime(slot, "%H:%M").time()
        slot_datetime = timezone.localize(datetime.combine(current_time.date(), slot_time))
        if slot_datetime > current_time:
            return slot_datetime
    # Если все слоты на сегодня уже прошли, возвращаем первый слот на завтра
    first_slot = datetime.strptime(today_slots[0], "%H:%M").time()
    next_day = current_time + timedelta(days=1)
    return timezone.localize(datetime.combine(next_day.date(), first_slot))

def send_telegram_message(bot, chat_id, message):
    """Отправляет сообщение в Telegram."""
    try:
        bot.send_message(chat_id, message)
        logging.info(f"Отправлено уведомление в Telegram")
    except Exception as e:
        logging.error(f"Не удалось отправить уведомление в Telegram: {str(e)}")

def check_product_notifications(product_info, min_rating, price_threshold, coupon_threshold):
    """Проверяет условия для отправки уведомлений."""
    notifications = []

    # Проверка рейтинга
    rating = product_info.get("Rating", "Not Found")
    if rating != "Not Found":
        try:
            rating_value = float(rating)
            if rating_value < min_rating:
                notifications.append(f"⚠️ Низкий рейтинг: {rating} звезд для ASIN {product_info['ASIN']}")
        except ValueError:
            notifications.append(f"⚠️ Некорректный рейтинг для ASIN {product_info['ASIN']}")

    # Проверка изменения цены
    full_price = product_info.get("Price", "Not Found")
    prime_price = product_info.get("Prime Price", "Not Found")
    if full_price != "Not Found" and prime_price != "Not Found":
        try:
            full_price_value = float(re.sub(r'[^\d.]', '', str(full_price))) if isinstance(full_price, (str, float)) else None
            prime_price_value = float(re.sub(r'[^\d.]', '', str(prime_price))) if isinstance(prime_price, (str, float)) else None
            if full_price_value and prime_price_value:
                price_diff_percent = abs(full_price_value - prime_price_value) / full_price_value * 100
                if price_diff_percent >= price_threshold:
                    notifications.append(f"💰 Значительное изменение цены для ASIN {product_info['ASIN']}: Full {full_price}, Prime {prime_price}")
        except ValueError:
            notifications.append(f"⚠️ Некорректная цена для ASIN {product_info['ASIN']}")

    # Проверка купона
    coupon_discount = product_info.get("Coupon Discount", "Not Found")
    if coupon_discount != "Not Found":
        try:
            coupon_value = float(re.sub(r'[^\d.]', '', str(coupon_discount).replace('%', '')))
            if coupon_value >= coupon_threshold:
                notifications.append(f"🏷️ Большой купон для ASIN {product_info['ASIN']}: {coupon_value}%")
        except ValueError:
            notifications.append(f"⚠️ Некорректная скидка купона для ASIN {product_info['ASIN']}")

    return notifications

def create_xlsx_report(data, current_time_str):
    """Создание XLSX отчета с данными о продуктах."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Product Report"

        # Стили
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        bold_blue_font = Font(bold=True, color="0000FF")  # Жирный и синий цвет для кликабельных ASIN

        # Заголовок
        ws['A1'] = f"Product Report - {current_time_str}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:N1')  # Расширил диапазон для новых колонок

        # Заголовки столбцов
        headers = [
            "Company", "ASIN", "Title", "Full Price", "Prime Price", 
            "List Price", "Sale Price", "Prime Price",
            "Rating", "Number of Reviews", "Coupon Discount", 
            "Final Price", "Discount Percent", "Variations Count"
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Данные
        row = 3
        for company, products in data.items():
            for product in products:
                ws.cell(row=row, column=1, value=company)
                asin_cell = ws.cell(row=row, column=2, value=product.get('ASIN', 'Не найдено'))
                asin = product.get('ASIN', 'Не найдено')
                if asin != 'Не найдено':
                    asin_cell.hyperlink = f"https://www.amazon.de/dp/{asin}"
                    asin_cell.font = bold_blue_font  # Сделать кликабельным ASIN жирным и синим
                ws.cell(row=row, column=3, value=product.get('Title', 'Не найдено'))
                ws.cell(row=row, column=4, value=product.get('Price', 'Не найдено'))
                ws.cell(row=row, column=5, value=product.get('Prime Price', 'Не найдено'))
                ws.cell(row=row, column=6, value=product.get('List Price', 'Not Found'))
                ws.cell(row=row, column=7, value=product.get('Sale Price', 'Not Found'))
                ws.cell(row=row, column=8, value=product.get('Prime Price', 'Not Found'))
                ws.cell(row=row, column=9, value=product.get('Rating', 'Не найдено'))
                ws.cell(row=row, column=10, value=product.get('Number of Reviews', 'Не найдено'))
                ws.cell(row=row, column=11, value=product.get('Coupon Discount', 'Not Found'))
                ws.cell(row=row, column=12, value=product.get('Final Price', 'Not Found'))
                ws.cell(row=row, column=13, value=product.get('Discount Percent', 'Not Found'))
                ws.cell(row=row, column=14, value=product.get('Variations Count', 'Not Found'))
                row += 1

        # Автоматическая регулировка ширины столбцов
        for col in ws.columns:
            max_length = 0
            column = col[0].column  # Получаем индекс столбца
            column_letter = get_column_letter(column)
            for cell in col:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Сохранение файла
        filename = f"product_report_{current_time_str.replace(':', '_')}.xlsx"
        wb.save(filename)
        return filename
    except Exception as e:
        logging.error(f"Ошибка при создании XLSX отчета: {str(e)}")
        return None

def send_telegram_notification(config, current_time_str, data):
    """Отправка уведомления и отчета в Telegram."""
    bot = telebot.TeleBot(config.get('telegram_bot_token', ''))
    chat_id = config.get('telegram_chat_id', '')
    if not chat_id:
        logging.error("Telegram chat_id не установлен в конфигурации.")
        return

    # Формирование сообщения
    message = f"Отчет за {current_time_str}\n\n"
    for company, products in data.items():
        message += f"{company}:\n"
        for product in products:
            message += f" ASIN: {product['ASIN']}\n"
            message += f" Цена: {product.get('Price', 'Не найдено')}\n"
            message += f" Рейтинг: {product.get('Rating', 'Не найдено')}\n\n"

    # Отправка сообщения
    send_telegram_message(bot, chat_id, message)

    # Создание и отправка XLSX файла
    xlsx_filename = create_xlsx_report(data, current_time_str)
    if xlsx_filename:
        try:
            with open(xlsx_filename, 'rb') as report_file:
                bot.send_document(chat_id, report_file)
            logging.info("Отчет успешно отправлен в Telegram.")
        except Exception as e:
            logging.error(f"Не удалось отправить отчет в Telegram: {str(e)}")
        finally:
            if os.path.exists(xlsx_filename):
                os.remove(xlsx_filename)

def apply_formatting(sheet, header, start_row, data_length):
    """
    Применяет форматирование к заголовкам и определенным ячейкам.
    Оптимизировано для уменьшения количества запросов к API.
    """
    # Форматирование заголовков
    header_format = CellFormat(
        backgroundColor=Color(0.85, 0.93, 0.83),  # Цвет #d9ead3
        textFormat=TextFormat(bold=True)
    )
    header_range = f"A1:{get_column_letter(len(header))}1"
    format_cell_range(sheet, header_range, header_format)

    # Получаем все значения в колонке "Наименование" (A) за один запрос
    end_row = start_row + data_length - 1
    name_column_range = f"A{start_row}:{get_column_letter(1)}{end_row}"
    name_values = sheet.get(name_column_range)

    # Подготовка списков для форматирования
    parent_variations_rows = []
    params_rows = []

    params_to_highlight = ["BSR", "Number of Reviews", "Rating", "Price", "List Price", "Prime Price"]

    for idx, row in enumerate(name_values, start=start_row):
        cell_value = row[0]
        if cell_value in ["Parent ASIN", "Variations ASIN"]:
            parent_variations_rows.append(idx)
        if cell_value in params_to_highlight:
            params_rows.append(idx)

    # Форматирование ячеек "Parent ASIN" и "Variations ASIN"
    if parent_variations_rows:
        for row in parent_variations_rows:
            cell_range = f"A{row}"
            format_cell_range(sheet, cell_range, CellFormat(
                backgroundColor=Color(1, 1, 0),  # Цвет #ffff00
                textFormat=TextFormat(bold=True)
            ))

    # Форматирование параметров
    if params_rows:
        for row in params_rows:
            cell_range = f"A{row}"
            format_cell_range(sheet, cell_range, CellFormat(
                backgroundColor=Color(0.85, 0.93, 0.83),  # Цвет #d9ead3
                textFormat=TextFormat(bold=True)
            ))

def find_nearest_slot(current_time_slot_formatted, all_slots):
    """Определяет ближайший временной слот к текущему времени."""
    try:
        current_time = datetime.strptime(current_time_slot_formatted, "%H:%M")
        slot_times = []
        for slot in all_slots:
            slot_time = datetime.strptime(slot, "%H:%M")
            slot_times.append(slot_time)

        # Находим ближайшее время
        nearest_time = min(slot_times, key=lambda x: abs((x - current_time).total_seconds()))
        nearest_slot = nearest_time.strftime("%H:%M")
        return nearest_slot
    except Exception as e:
        logging.error(f"Ошибка при определении ближайшего временного слота: {e}")
        return None


def find_credentials_file():
    """Пытается найти файл учетных данных по нескольким возможным путям."""
    possible_paths = [
        os.path.join(os.path.expanduser('~'), 'Downloads', 'maximumstores53-24d4ef8c1298.json'),
        os.path.join(os.getcwd(), 'maximumstores53-24d4ef8c1298.json'),
        'maximumstores53-24d4ef8c1298.json'
    ]
    
    for path in possible_paths:
        if os.path.exists(path):
            logging.info(f"Найден файл учетных данных: {path}")
            return path

    logging.error("Файл учетных данных не найден ни в одном из возможных путей.")
    return None

def get_product_data(api_key, asin, domain='de'):
    """
    Получает данные о продукте Amazon через API ScrapingDog.
    
    :param api_key: API ключ ScrapingDog.
    :param asin: ASIN продукта.
    :param domain: Домен Amazon (по умолчанию 'de' для Amazon.de).
    :return: JSON-ответ или None в случае ошибки.
    """
    # Определяем страну на основе домена
    domain_to_country = {
        'com': 'us',
        'de': 'de',
        'co.uk': 'uk',
        'fr': 'fr',
        'it': 'it',
        'es': 'es',
        'ca': 'ca',
        'co.jp': 'jp',
        'com.au': 'au',
        'nl': 'nl',
        'se': 'se',
        'sg': 'sg',
        'in': 'in',
        'com.br': 'br',
        'ae': 'ae',
        # Добавьте другие домены и соответствующие страны по необходимости
    }
    
    country = domain_to_country.get(domain, 'us')  # По умолчанию 'us'
    
    url = "https://api.scrapingdog.com/amazon/product"
    
    params = {
        "api_key": api_key,
        "asin": asin,
        "domain": domain,
        "country": country  # Добавляем параметр country
    }
    
    try:
        api_limiter.wait()  # Ждем, чтобы не превысить лимит запросов
        response = requests.get(url, params=params, timeout=30)
        if response.status_code == 200:
            data = response.json()
            logging.debug(f"Получены данные от ScrapingDog для ASIN {asin}: {json.dumps(data, indent=2, ensure_ascii=False)}")
            return data
        else:
            logging.error(f"Запрос не удался с кодом статуса: {response.status_code}")
            logging.error(f"Содержимое ответа: {response.text}")
            return None
    except requests.exceptions.RequestException as e:
        logging.error(f"Ошибка при запросе к ScrapingDog для ASIN {asin}: {str(e)}")
        return None
def gather_product_data(config):
    """Функция для сбора данных по продуктам. Возвращает текущие результаты."""
    current_results = {}

    # Получаем имя нашей компании из конфига или используем значение по умолчанию
    our_company_name = config.get('company_name', 'Merino.tech. (Мы)')
    current_results[our_company_name] = []

    # API ключ ScrapingDog
    scrapingdog_api_key = config.get('ScrapingDogAPIKey', '').strip()

    if not scrapingdog_api_key:
        logging.error("API ключ ScrapingDog не найден в конфигурации.")
        return current_results

    # Обработка основных продуктов (Parent ASIN) нашей компании
    logging.info(f"Начинаем сбор данных по основным продуктам {our_company_name} (Parent ASIN).")
    for url in config.get('product_urls', []):
        logging.debug(f"Обработка Parent ASIN по URL: {url}")
        try:
            asin = extract_asin(url)
            if asin == 'Not Found':
                logging.warning(f"ASIN не найден для URL: {url}")
                continue

            # Определение домена из URL
            parsed_url = urlparse(url)
            domain_parts = parsed_url.netloc.split('.')
            if len(domain_parts) >= 2:
                domain = domain_parts[-1]  # Например, 'de' из 'amazon.de'
            else:
                domain = 'com'  # По умолчанию

            product_data = get_product_data(scrapingdog_api_key, asin, domain=domain)
            if product_data:
                # Преобразование данных из ScrapingDog в формат, используемый в скрипте
                currency_code = determine_currency(url)
                currency_symbol = CURRENCY_SYMBOLS.get(currency_code, '$')

                price = extract_price(product_data.get('price'), currency_code)
                # ЗАМЕНА list_price на previous_price
                list_price = extract_price(product_data.get('previous_price'), currency_code)
                prime_price = price  # Если у вас есть отдельное поле для Prime Price, используйте его
                coupon = extract_coupon(product_data.get('coupon_text'))  # Изменено на 'coupon_text'

                final_price = calculate_final_price(price, prime_price, coupon, currency_symbol)
                discount_percent = calculate_discount_percent(price, final_price)

                rating = extract_rating(product_data.get('average_rating'))
                total_reviews_data = product_data.get('total_reviews', 'Not Found')
                reviews_count = extract_reviews_count(total_reviews_data)
                product_information = product_data.get('product_information', {})
                bsr = extract_bsr(product_information)  # Передаём весь словарь
                brand = product_data.get('brand', 'Not Found')

                # Извлечение информации о Prime Exclusive
                is_prime_exclusive = product_data.get('is_prime_exclusive', False)
                if isinstance(is_prime_exclusive, str):
                    is_prime_exclusive = is_prime_exclusive.lower() == 'true'
                prime_exclusive_message = product_data.get('prime_exclusive_message', '')

                # Очистка сообщения от JavaScript-кода
                prime_exclusive_message_clean = re.split(r'\(function', prime_exclusive_message)[0].strip()

                # Если продукт является Prime Exclusive, извлекаем Prime Price из сообщения
                if is_prime_exclusive and prime_exclusive_message_clean:
                    extracted_prime_price = extract_prime_price_from_message(prime_exclusive_message_clean)
                    if extracted_prime_price != "Not Found":
                        prime_price = extracted_prime_price
                        # Пересчитываем итоговую цену и процент скидки с новым Prime Price
                        final_price = calculate_final_price(price, float(prime_price.replace(',', '.').replace(' €', '')), coupon, currency_symbol)
                        discount_percent = calculate_discount_percent(price, final_price)
                    else:
                        logging.warning(f"Не удалось извлечь Prime Price из сообщения для ASIN {asin}")

                product_info = {
                    "ASIN": asin,
                    "Title": product_data.get('title', 'Не найдено'),
                    "Price": price,
                    "Prime Price": prime_price,
                    "List Price": list_price,
                    "Coupon Discount": coupon,
                    "Final Price": final_price,
                    "Discount Percent": discount_percent,
                    "Rating": rating,
                    "Number of Reviews": reviews_count,
                    "BSR": bsr,
                    "Brand": brand,
                    "Scrape Date": get_kyiv_time().strftime("%d.%m.%Y"),
                    "URL": url,
                    "is_prime_exclusive": is_prime_exclusive,
                    "prime_exclusive_message": prime_exclusive_message_clean
                }

                # Добавляем логирование извлечённых данных
                logging.info(f"Извлеченные данные для ASIN {asin}:")
                for key, value in product_info.items():
                    logging.info(f"  {key}: {value}")

                current_results[our_company_name].append(product_info)
                logging.info(f"Успешно собраны данные для ASIN {asin}")
            else:
                logging.warning(f"Не удалось получить данные для ASIN {asin}")

        except Exception as e:
            logging.error(f"Ошибка при сборе данных для ASIN {asin}: {str(e)}")

    # Обработка вариаций продуктов (Variation ASIN) вашей компании
    logging.info(f"Начинаем сбор данных по вариациям продуктов {our_company_name} (Variation ASIN).")
    for var_url in config.get('variation_urls', []):
        logging.debug(f"Обработка Variation ASIN по URL: {var_url}")
        try:
            asin = extract_asin(var_url)
            if asin == 'Not Found':
                logging.warning(f"ASIN не найден для URL: {var_url}")
                continue

            # Определение домена из URL
            parsed_url = urlparse(var_url)
            domain_parts = parsed_url.netloc.split('.')
            if len(domain_parts) >= 2:
                domain = domain_parts[-1]
            else:
                domain = 'com'

            product_data = get_product_data(scrapingdog_api_key, asin, domain=domain)
            if product_data:
                currency_code = determine_currency(var_url)
                currency_symbol = CURRENCY_SYMBOLS.get(currency_code, '$')

                price = extract_price(product_data.get('price'), currency_code)
                # ЗАМЕНА list_price на previous_price
                list_price = extract_price(product_data.get('previous_price'), currency_code)
                prime_price = price
                coupon = extract_coupon(product_data.get('coupon_text'))  # Изменено на 'coupon_text'

                final_price = calculate_final_price(price, prime_price, coupon, currency_symbol)
                discount_percent = calculate_discount_percent(price, final_price)

                rating = extract_rating(product_data.get('average_rating'))
                total_reviews_data = product_data.get('total_reviews', 'Not Found')
                reviews_count = extract_reviews_count(total_reviews_data)
                product_information = product_data.get('product_information', {})
                bsr = extract_bsr(product_information)  # Передаём весь словарь
                brand = product_data.get('brand', 'Not Found')

                # Извлечение информации о Prime Exclusive
                is_prime_exclusive = product_data.get('is_prime_exclusive', False)
                if isinstance(is_prime_exclusive, str):
                    is_prime_exclusive = is_prime_exclusive.lower() == 'true'
                prime_exclusive_message = product_data.get('prime_exclusive_message', '')

                # Очистка сообщения от JavaScript-кода
                prime_exclusive_message_clean = re.split(r'\(function', prime_exclusive_message)[0].strip()

                # Если продукт является Prime Exclusive, извлекаем Prime Price из сообщения
                if is_prime_exclusive and prime_exclusive_message_clean:
                    extracted_prime_price = extract_prime_price_from_message(prime_exclusive_message_clean)
                    if extracted_prime_price != "Not Found":
                        prime_price = extracted_prime_price
                        # Пересчитываем итоговую цену и процент скидки с новым Prime Price
                        final_price = calculate_final_price(price, float(prime_price.replace(',', '.').replace(' €', '')), coupon, currency_symbol)
                        discount_percent = calculate_discount_percent(price, final_price)
                    else:
                        logging.warning(f"Не удалось извлечь Prime Price из сообщения для ASIN {asin}")

                product_info = {
                    "ASIN": asin,
                    "Title": product_data.get('title', 'Не найдено'),
                    "Price": price,
                    "Prime Price": prime_price,
                    "List Price": list_price,
                    "Coupon Discount": coupon,
                    "Final Price": final_price,
                    "Discount Percent": discount_percent,
                    "Rating": rating,
                    "Number of Reviews": reviews_count,
                    "BSR": bsr,
                    "Brand": brand,
                    "Scrape Date": get_kyiv_time().strftime("%d.%m.%Y"),
                    "URL": var_url,
                    "is_prime_exclusive": is_prime_exclusive,
                    "prime_exclusive_message": prime_exclusive_message_clean
                }

                # Добавляем логирование извлечённых данных
                logging.info(f"Извлеченные данные для вариации ASIN {asin}:")
                for key, value in product_info.items():
                    logging.info(f"  {key}: {value}")

                current_results[our_company_name].append(product_info)
                logging.info(f"Успешно собраны данные для вариации ASIN {asin}")
            else:
                logging.warning(f"Не удалось получить данные для вариации ASIN {asin}")

        except Exception as e:
            logging.error(f"Ошибка при сборе данных для вариации ASIN {asin}: {str(e)}")

    # Обработка конкурентов
    logging.info("Начинаем сбор данных по конкурентам.")
    competitor_names = config.get('competitor_names', {})
    for i in range(1, 6):
        competitor_name = competitor_names.get(str(i))
        if competitor_name:
            competitor_urls_key = f'{i}competitor_urls'
            competitor_variation_urls_key = f'{i}variation_urls'

            competitor_urls = config.get(competitor_urls_key, [])
            competitor_variation_urls = config.get(competitor_variation_urls_key, [])

            current_results[competitor_name] = []

            # Обработка основных продуктов конкурента (Parent ASIN)
            logging.info(f"Обработка основных продуктов конкурента {competitor_name} (Parent ASIN).")
            for url in competitor_urls:
                logging.debug(f"Обработка продукта конкурента {competitor_name} по URL: {url}")
                try:
                    asin = extract_asin(url)
                    if asin == 'Not Found':
                        logging.warning(f"ASIN не найден для URL: {url}")
                        continue

                    parsed_url = urlparse(url)
                    domain_parts = parsed_url.netloc.split('.')
                    if len(domain_parts) >= 2:
                        domain = domain_parts[-1]
                    else:
                        domain = 'com'

                    product_data = get_product_data(scrapingdog_api_key, asin, domain=domain)
                    if product_data:
                        currency_code = determine_currency(url)
                        currency_symbol = CURRENCY_SYMBOLS.get(currency_code, '$')

                        price = extract_price(product_data.get('price'), currency_code)
                        # ЗАМЕНА list_price на previous_price
                        list_price = extract_price(product_data.get('previous_price'), currency_code)
                        prime_price = price
                        coupon = extract_coupon(product_data.get('coupon_text'))  # Изменено на 'coupon_text'

                        final_price = calculate_final_price(price, prime_price, coupon, currency_symbol)
                        discount_percent = calculate_discount_percent(price, final_price)

                        rating = extract_rating(product_data.get('average_rating'))
                        total_reviews_data = product_data.get('total_reviews', 'Not Found')
                        reviews_count = extract_reviews_count(total_reviews_data)
                        product_information = product_data.get('product_information', {})
                        bsr = extract_bsr(product_information)  # Передаём весь словарь
                        brand = product_data.get('brand', 'Not Found')

                        # Извлечение информации о Prime Exclusive
                        is_prime_exclusive = product_data.get('is_prime_exclusive', False)
                        if isinstance(is_prime_exclusive, str):
                            is_prime_exclusive = is_prime_exclusive.lower() == 'true'
                        prime_exclusive_message = product_data.get('prime_exclusive_message', '')

                        # Очистка сообщения от JavaScript-кода
                        prime_exclusive_message_clean = re.split(r'\(function', prime_exclusive_message)[0].strip()

                        # Если продукт является Prime Exclusive, извлекаем Prime Price из сообщения
                        if is_prime_exclusive and prime_exclusive_message_clean:
                            extracted_prime_price = extract_prime_price_from_message(prime_exclusive_message_clean)
                            if extracted_prime_price != "Not Found":
                                prime_price = extracted_prime_price
                                # Пересчитываем итоговую цену и процент скидки с новым Prime Price
                                final_price = calculate_final_price(price, float(prime_price.replace(',', '.').replace(' €', '')), coupon, currency_symbol)
                                discount_percent = calculate_discount_percent(price, final_price)
                            else:
                                logging.warning(f"Не удалось извлечь Prime Price из сообщения для ASIN {asin}")

                        product_info = {
                            "ASIN": asin,
                            "Title": product_data.get('title', 'Не найдено'),
                            "Price": price,
                            "Prime Price": prime_price,
                            "List Price": list_price,
                            "Coupon Discount": coupon,
                            "Final Price": final_price,
                            "Discount Percent": discount_percent,
                            "Rating": rating,
                            "Number of Reviews": reviews_count,
                            "BSR": bsr,
                            "Brand": brand,
                            "Scrape Date": get_kyiv_time().strftime("%d.%m.%Y"),
                            "URL": url,
                            "is_prime_exclusive": is_prime_exclusive,
                            "prime_exclusive_message": prime_exclusive_message_clean
                        }

                        # Добавляем логирование извлечённых данных
                        logging.info(f"Извлеченные данные для конкурента {competitor_name}, ASIN {asin}:")
                        for key, value in product_info.items():
                            logging.info(f"  {key}: {value}")

                        current_results[competitor_name].append(product_info)
                        logging.info(f"Успешно собраны данные для конкурента {competitor_name}: ASIN {asin}")
                    else:
                        logging.warning(f"Не удалось получить данные для конкурента {competitor_name}: ASIN {asin}")

                except Exception as e:
                    logging.error(f"Ошибка при сборе данных для конкурента {competitor_name} (ASIN {asin}): {str(e)}")

            # Обработка вариаций продуктов конкурента (Variation ASIN)
            logging.info(f"Обработка вариаций продуктов конкурента {competitor_name} (Variation ASIN).")
            for var_url in competitor_variation_urls:
                logging.debug(f"Обработка вариации продукта конкурента {competitor_name} по URL: {var_url}")
                try:
                    asin = extract_asin(var_url)
                    if asin == 'Not Found':
                        logging.warning(f"ASIN не найден для URL: {var_url}")
                        continue

                    parsed_url = urlparse(var_url)
                    domain_parts = parsed_url.netloc.split('.')
                    if len(domain_parts) >= 2:
                        domain = domain_parts[-1]
                    else:
                        domain = 'com'

                    product_data = get_product_data(scrapingdog_api_key, asin, domain=domain)
                    if product_data:
                        currency_code = determine_currency(var_url)
                        currency_symbol = CURRENCY_SYMBOLS.get(currency_code, '$')

                        price = extract_price(product_data.get('price'), currency_code)
                        # ЗАМЕНА list_price на previous_price
                        list_price = extract_price(product_data.get('previous_price'), currency_code)
                        prime_price = price
                        coupon = extract_coupon(product_data.get('coupon_text'))  # Изменено на 'coupon_text'

                        final_price = calculate_final_price(price, prime_price, coupon, currency_symbol)
                        discount_percent = calculate_discount_percent(price, final_price)

                        rating = extract_rating(product_data.get('average_rating'))
                        total_reviews_data = product_data.get('total_reviews', 'Not Found')
                        reviews_count = extract_reviews_count(total_reviews_data)
                        product_information = product_data.get('product_information', {})
                        bsr = extract_bsr(product_information)  # Передаём весь словарь
                        brand = product_data.get('brand', 'Not Found')

                        # Извлечение информации о Prime Exclusive
                        is_prime_exclusive = product_data.get('is_prime_exclusive', False)
                        if isinstance(is_prime_exclusive, str):
                            is_prime_exclusive = is_prime_exclusive.lower() == 'true'
                        prime_exclusive_message = product_data.get('prime_exclusive_message', '')

                        # Очистка сообщения от JavaScript-кода
                        prime_exclusive_message_clean = re.split(r'\(function', prime_exclusive_message)[0].strip()

                        # Если продукт является Prime Exclusive, извлекаем Prime Price из сообщения
                        if is_prime_exclusive and prime_exclusive_message_clean:
                            extracted_prime_price = extract_prime_price_from_message(prime_exclusive_message_clean)
                            if extracted_prime_price != "Not Found":
                                prime_price = extracted_prime_price
                                # Пересчитываем итоговую цену и процент скидки с новым Prime Price
                                final_price = calculate_final_price(price, float(prime_price.replace(',', '.').replace(' €', '')), coupon, currency_symbol)
                                discount_percent = calculate_discount_percent(price, final_price)
                            else:
                                logging.warning(f"Не удалось извлечь Prime Price из сообщения для ASIN {asin}")

                        product_info = {
                            "ASIN": asin,
                            "Title": product_data.get('title', 'Не найдено'),
                            "Price": price,
                            "Prime Price": prime_price,
                            "List Price": list_price,
                            "Coupon Discount": coupon,
                            "Final Price": final_price,
                            "Discount Percent": discount_percent,
                            "Rating": rating,
                            "Number of Reviews": reviews_count,
                            "BSR": bsr,
                            "Brand": brand,
                            "Scrape Date": get_kyiv_time().strftime("%d.%m.%Y"),
                            "URL": var_url,
                            "is_prime_exclusive": is_prime_exclusive,
                            "prime_exclusive_message": prime_exclusive_message_clean
                        }

                        # Добавляем логирование извлечённых данных
                        logging.info(f"Извлеченные данные для вариации конкурента {competitor_name}, ASIN {asin}:")
                        for key, value in product_info.items():
                            logging.info(f"  {key}: {value}")

                        current_results[competitor_name].append(product_info)
                        logging.info(f"Успешно собраны данные для вариации конкурента {competitor_name}: ASIN {asin}")
                    else:
                        logging.warning(f"Не удалось получить данные для вариации конкурента {competitor_name}: ASIN {asin}")

                except Exception as e:
                    logging.error(f"Ошибка при сборе данных для вариации конкурента {competitor_name} (ASIN {asin}): {str(e)}")

    return current_results
def extract_prime_price_from_message(message):
    """
    Извлекает цену из сообщения prime_exclusive_message.
    
    :param message: Строка с сообщением Prime Exclusive.
    :return: Цена как строка, например "43,19 €" или "Not Found".
    """
    try:
        # Регулярное выражение для поиска цены в формате "XX,XX €"
        match = re.search(r'kauf(?:e|en)?(?: diesen Artikel)? bei (\d+[.,]\d{2})\s*€', message, re.IGNORECASE)
        if match:
            price_str = match.group(1).replace(',', '.')
            price_float = float(price_str)
            return f"{price_float:.2f} €"
    except Exception as e:
        logging.error(f"Ошибка при извлечении Prime Price из сообщения: {e}")
    return "Not Found"


def extract_bsr(product_information):
    """Извлекает BSR (Best Sellers Rank) из product_information."""
    bsr_keys = ['Amazon Bestseller-Rang', 'Amazon BestsellerRang', 'Best Sellers Rank']
    for key in bsr_keys:
        bsr_data = product_information.get(key, None)
        if bsr_data:
            logging.debug(f"Найдено BSR по ключу '{key}': {bsr_data}")
            # Извлечение числа после 'Nr.' или аналогичного обозначения
            bsr_match = re.search(r'Nr\.\s?([\d\.]+)', bsr_data)
            if bsr_match:
                rank_str = bsr_match.group(1)
                # Удаляем точки, используемые как разделители тысяч (например, '2.366' -> '2366')
                rank_str = rank_str.replace('.', '')
                try:
                    rank_int = int(rank_str)
                    return rank_int
                except ValueError:
                    logging.error(f"Не удалось преобразовать BSR '{rank_str}' в число.")
    logging.warning("BSR не найден в product_information.")
    return 'Not Found'

def extract_reviews_count(reviews_data):
    """Извлекает количество отзывов из данных."""
    logging.debug(f"Извлечение количества отзывов из данных: {reviews_data}")
    if isinstance(reviews_data, (int, float)):
        return int(reviews_data)
    elif isinstance(reviews_data, str):
        # Ищем число в строке
        reviews_match = re.search(r'(\d[\d,\.]*)', reviews_data)
        if reviews_match:
            reviews_str = reviews_match.group(1)
            # Удаляем запятые и точки для получения целого числа
            reviews_str = reviews_str.replace(',', '').replace('.', '')
            try:
                reviews_int = int(reviews_str)
                return reviews_int
            except ValueError:
                logging.error(f"Не удалось преобразовать количество отзывов '{reviews_str}' в число.")
    return 'Not Found'
def update_monitoring_sheet(spreadsheet, data, current_time_slot, config, sheet_name):
    """
    Обновляет данные на указанном листе Google Sheets и применяет форматирование.
    """
    try:
        sheet = spreadsheet.worksheet(sheet_name)
    except gspread.exceptions.WorksheetNotFound:
        logging.error(f"Лист '{sheet_name}' не найден в таблице.")
        return

    # Формируем заголовки для колонок
    header = [
        "Наименование", "Параметры", "Данные"
    ] + config.get('active_trade_slots', []) + ["Зона анализа"] + config.get('analysis_slots', [])

    header = [slot.strip() if isinstance(slot, str) else slot for slot in header]

    logging.debug(f"Текущий временной слот: '{current_time_slot}'")
    logging.debug(f"Заголовки таблицы: {header}")

    value_ranges = [{
        'range': f'{sheet_name}!A1',
        'values': [header]
    }]

    start_row = 3
    current_row = start_row

    # Получаем текущее время в формате "YYYY-MM-DD HH:MM:SS" в указанном часовом поясе
    current_time = get_kyiv_time(config.get('timezone', 'Europe/Kiev'))
    current_time_formatted = current_time.strftime("%Y-%m-%d %H:%M:%S")
    current_time_slot_formatted = current_time.strftime("%H:%M")

    # Добавляем запись текущего времени на 2-ю строку в колонку "Данные"
    time_notation = f'{sheet_name}!C2'  # Ячейка C2 (строка 2, колонка "Данные")
    value_ranges.append({
        'range': time_notation,
        'values': [[current_time_formatted]]  # Записываем текущее время
    })

    parameters = ["BSR", "Number of Reviews", "Rating", "Price"]

    # Получаем имя нашей компании
    our_company_name = config.get('company_name', 'Merino.tech. (Мы)')

    # Получаем имена конкурентов из конфига
    competitor_names = config.get('competitor_names', {})
    companies = [(section, our_company_name) for section in ['product_urls']]
    for i in range(1, 6):
        competitor_name = competitor_names.get(str(i))
        if competitor_name:
            companies.append((f'{i}competitor_urls', competitor_name))

    data_to_write = []
    
    # Безопасное получение значения
    def safe_get_value(product_info, key):
        value = product_info.get(key)
        if value is None or str(value).lower() == 'not found':
            return 0
        return value

    # Форматирование значения
    def format_value(value):
        if isinstance(value, (int, float)):
            return f"{value:.2f}"  # Убрал знак $
        return str(value) if value is not None else '0'

    data_to_write.append(["Parent ASIN"])
    current_row += 1

    asin_row_mapping_parent = {}

    # Parent ASIN обработка
    for param in parameters:
        data_to_write.append([param])
        current_row += 1

        for section, company_name in companies:
            urls = config.get(section, [])
            if urls:
                for url in urls:
                    asin = extract_asin(url)
                    logging.debug(f"Обработка URL: {url}, извлеченный ASIN: {asin}, компания: {company_name}")

                    product_info = next((prod for prod in data.get(company_name, []) if prod.get('ASIN') == asin), {})

                    # Создание гиперссылки
                    hyperlink_formula = f'=HYPERLINK("{url}", "{asin}")'

                    # Безопасное получение и форматирование значения
                    value = safe_get_value(product_info, param)
                    formatted_value = format_value(value)

                    data_to_write.append([company_name, hyperlink_formula, formatted_value])
                    asin_row_mapping_parent[(company_name, asin, param)] = current_row
                    current_row += 1

    data_to_write.append(["Variations ASIN"])
    current_row += 1

    variations_sections = [(section, our_company_name) for section in ['variation_urls']]
    for i in range(1, 6):
        competitor_name = competitor_names.get(str(i))
        if competitor_name:
            variations_sections.append((f'{i}variation_urls', competitor_name))

    price_types = ["Price", "List Price", "Prime Price"]
    asin_row_mapping_variations = {}
    average_row_mapping = {}

    # Variations ASIN обработка
    for price_type in price_types:
        data_to_write.append([price_type])
        current_row += 1

        for section, company_name in variations_sections:
            urls = config.get(section, [])
            if urls:
                temp_asin_rows = []
                num_variations = 0

                formula_row = current_row

                # Добавляем строку для средней цены в колонку "Данные"
                data_to_write.append([company_name, "Средняя цена", ""])
                current_row += 1

                # Сохраняем номер строки для формулы средней цены
                average_row_number = current_row - 1
                average_row_mapping[(company_name, price_type)] = average_row_number

                for url in urls:
                    asin = extract_asin(url)
                    logging.debug(f"Обработка вариации URL: {url}, извлеченный ASIN: {asin}, компания: {company_name}")

                    product_info = next((prod for prod in data.get(company_name, []) if prod.get('ASIN') == asin), {})

                    # Безопасное получение цены
                    price_value = safe_get_value(product_info, price_type)
                    price_display = format_value(price_value)

                    # Создание гиперссылки
                    hyperlink_formula = f'=HYPERLINK("{url}", "{asin}")'

                    temp_asin_rows.append([company_name, hyperlink_formula, price_display])
                    asin_row_mapping_variations[(company_name, asin, price_type)] = current_row
                    current_row += 1
                    num_variations += 1

                data_to_write.extend(temp_asin_rows)

                # Обновляем формулу средней цены в колонке "Данные"
                first_price_row = formula_row + 1
                last_price_row = formula_row + num_variations
                price_range = f"C{first_price_row}:C{last_price_row}"
                average_formula = f'=AVERAGE(FILTER({price_range}, {price_range}<>""))'
                data_to_write[formula_row - start_row][2] = average_formula

    end_row = start_row + len(data_to_write) - 1
    range_notation = f'{sheet_name}!A{start_row}:C{end_row}'

    value_ranges.append({
        'range': range_notation,
        'values': data_to_write
    })

    # Определяем колонку 'Данные' (C)
    data_column_index = header.index("Данные") + 1
    data_column_letter = get_column_letter(data_column_index)

    # Определяем ближайший временной слот
    all_slots = config.get('active_trade_slots', []) + config.get('analysis_slots', [])
    nearest_slot = find_nearest_slot(current_time_slot_formatted, all_slots)
    if nearest_slot and nearest_slot in header:
        slot_column_index = header.index(nearest_slot) + 1
        slot_column_letter = get_column_letter(slot_column_index)
        logging.info(f"Данные будут записаны также в ближайший временной слот '{nearest_slot}' (столбец {slot_column_index})")
    else:
        slot_column_index = None
        slot_column_letter = None
        logging.warning("Не удалось определить ближайший временной слот. Данные будут записаны только в колонку 'Данные'")

    slot_updates = []

    def find_product_info(all_data, company_name, asin):
        products = all_data.get(company_name, [])
        for prod in products:
            if prod.get('ASIN') == asin:
                return prod
        return None

    # Обновление данных для Parent ASIN
    for (company_name, asin, param), row_number in asin_row_mapping_parent.items():
        product_info = find_product_info(data, company_name, asin)
        if product_info:
            value = safe_get_value(product_info, param)
            formatted_value = format_value(value)

            # Записываем в колонку 'Данные'
            data_cell_notation = f'{sheet_name}!{data_column_letter}{row_number}'
            slot_updates.append({
                'range': data_cell_notation,
                'values': [[formatted_value]]
            })

            # Записываем в ближайший временной слот
            if slot_column_letter:
                slot_cell_notation = f'{sheet_name}!{slot_column_letter}{row_number}'
                slot_updates.append({
                    'range': slot_cell_notation,
                    'values': [[formatted_value]]
                })

    # Обновление данных для Variations ASIN
    for (company_name, asin, price_type), row_number in asin_row_mapping_variations.items():
        product_info = find_product_info(data, company_name, asin)
        if product_info:
            price_value = safe_get_value(product_info, price_type)
            formatted_price = format_value(price_value)

            # Записываем в колонку 'Данные'
            data_cell_notation = f'{sheet_name}!{data_column_letter}{row_number}'
            slot_updates.append({
                'range': data_cell_notation,
                'values': [[formatted_price]]
            })

            # Записываем в ближайший временной слот
            if slot_column_letter:
                slot_cell_notation = f'{sheet_name}!{slot_column_letter}{row_number}'
                slot_updates.append({
                    'range': slot_cell_notation,
                    'values': [[formatted_price]]
                })

    # Обновление формулы средней цены
    for (company_name, price_type), row_number in average_row_mapping.items():
        variation_rows = [row_num for (comp, asin, ptype), row_num in asin_row_mapping_variations.items()
                          if comp == company_name and ptype == price_type]

        if variation_rows:
            first_row = min(variation_rows)
            last_row = max(variation_rows)

            # Формула для колонки 'Данные'
            data_price_range = f'{data_column_letter}{first_row}:{data_column_letter}{last_row}'
            average_formula_data = f'=AVERAGE(FILTER({data_price_range}, {data_price_range}<>""))'
            data_cell_notation = f'{sheet_name}!{data_column_letter}{row_number}'
            slot_updates.append({
                'range': data_cell_notation,
                'values': [[average_formula_data]]
            })

            # Формула для ближайшего временного слота
            if slot_column_letter:
                slot_price_range = f'{slot_column_letter}{first_row}:{slot_column_letter}{last_row}'
                average_formula_slot = f'=AVERAGE(FILTER({slot_price_range}, {slot_price_range}<>""))'
                slot_cell_notation = f'{sheet_name}!{slot_column_letter}{row_number}'
                slot_updates.append({
                    'range': slot_cell_notation,
                    'values': [[average_formula_slot]]
                })

    value_ranges.extend(slot_updates)

    try:
        data_body = {
            'value_input_option': 'USER_ENTERED',
            'data': value_ranges
        }
        spreadsheet.values_batch_update(data_body)
        logging.info(f"Данные успешно обновлены в листе '{sheet_name}' Google Sheets.")

        # Применение форматирования
        apply_formatting(sheet, header, start_row, len(data_to_write))

    except APIError as e:
        logging.error(f"Ошибка API при обновлении Google Sheets: {str(e)}")
        retry_delay = 60
        logging.info(f"Попытка повторного обновления через {retry_delay} секунд...")
        time.sleep(retry_delay)
        try:
            spreadsheet.values_batch_update(data_body)
            logging.info(f"Данные успешно обновлены в листе '{sheet_name}' Google Sheets после повторной попытки.")
        except Exception as e2:
            logging.error(f"Не удалось обновить Google Sheets после повторной попытки: {str(e2)}")
    except Exception as e:
        logging.error(f"Ошибка при обновлении Google Sheets: {str(e)}")



def update_google_sheets(current_results, spreadsheet_id, config, sheet_name, credentials_file):
    """Обновление Google Sheets данными из current_results."""
    try:
        client = authorize_google_sheets(credentials_file)
        spreadsheet = client.open_by_key(spreadsheet_id)
        
        # Получаем текущее время в формате HH:MM
        current_time_formatted = get_kyiv_time(config.get('timezone', 'Europe/Kiev')).strftime('%H:%M')
        
        all_slots = config.get('active_trade_slots', []) + config.get('analysis_slots', [])
        
        if current_time_formatted in all_slots:
            current_time_slot = current_time_formatted
        else:
            current_time_slot = None  # Текущее время не совпадает с временными слотами

        update_monitoring_sheet(
            spreadsheet,
            current_results,
            current_time_slot,
            config,
            sheet_name
        )

    except APIError as e:
        logging.error(f"Ошибка API при обновлении Google Sheets: {str(e)}")

def main():
    """Основная функция скрипта."""
    # Поиск файла учетных данных Google Sheets
    credentials_file = find_credentials_file()
    if not credentials_file:
        logging.critical("Не удалось найти файл учетных данных.")
        return

    # ID таблицы Google Sheets
    spreadsheet_id = '1ibuYnN9WeRZdHUqoiU2jFLez59fm5Gfgzeyvq7M4EaI'

    # Авторизация и загрузка основного конфига
    try:
        client = authorize_google_sheets(credentials_file)
        main_config = load_config_from_sheets(client, spreadsheet_id)
    except Exception as e:
        logging.critical(f"Не удалось авторизоваться или загрузить основной конфиг: {e}")
        return

    # Извлечение соответствий между конфигурационными листами и листами данных
    config_sheet_mappings = []
    for i in range(1, 10):  # Предполагаем, что у вас может быть до 10 листов, измените при необходимости
        config_key = f'Config_{i}'
        list_key = f'Name list_{i}'

        config_sheet_name = main_config.get(config_key)
        data_sheet_name = main_config.get(list_key)

        if config_sheet_name and data_sheet_name:
            config_sheet_mappings.append((config_sheet_name, data_sheet_name))
        else:
            continue  # Если одно из значений отсутствует, пропускаем

    if not config_sheet_mappings:
        logging.error("Не найдены конфигурационные листы в основном конфиге.")
        return

    # Получение временных слотов из конфигурации
    active_trade_slots = main_config.get('active_trade_slots', [])
    analysis_slots = main_config.get('analysis_slots', [])

    # Добавляем специальное время обновления, если оно задано
    update_time_slot = f"{main_config.get('update_time_hour', 0):02d}:{main_config.get('update_time_minute', 0):02d}"
    all_slots = active_trade_slots + analysis_slots + [update_time_slot]

    # Удаляем дубликаты и сортируем
    all_slots = sorted(list(set(all_slots)))

    logging.info(f"Все временные слоты: {all_slots}")

    def run_tasks():
        """Выполняет сбор данных и обновление для каждого листа."""
        for config_sheet_name, data_sheet_name in config_sheet_mappings:
            try:
                # Загрузка конфига для листа
                per_sheet_config = load_config_from_sheets(client, spreadsheet_id, config_sheet_name)

                # Объединяем основной конфиг и конфиг листа
                config = main_config.copy()
                config.update(per_sheet_config)

                # Выполнение сбора данных
                current_results = gather_product_data(config)

                # Обновление Google Sheets
                update_google_sheets(
                    current_results,
                    spreadsheet_id,
                    config,
                    data_sheet_name,
                    credentials_file
                )

                # Отправка уведомлений в Telegram
                if current_results:
                    current_time_slot = get_kyiv_time(config.get('timezone', 'Europe/Kiev')).strftime('%Y-%m-%d %H:%M:%S')
                    send_telegram_notification(
                        config,
                        current_time_slot,
                        current_results
                    )

                logging.info(f"Обработка листа '{data_sheet_name}' успешно завершена.")

            except Exception as e:
                logging.error(f"Ошибка при обработке листа '{data_sheet_name}': {e}")

    # **Выполняем задачи сразу при запуске скрипта**
    run_tasks()

    # Цикл регулярных запусков
    while True:
        try:
            # Используем main_config для получения timezone
            timezone_str = main_config.get('timezone', 'Europe/Kiev')
            current_time = get_kyiv_time(timezone_str)
            logging.info(f"Текущее время: {current_time.strftime('%Y-%m-%d %H:%M:%S')}")

            # Определение ближайшего следующего слота
            next_slot_time = get_next_slot(current_time, all_slots, timezone_str)
            logging.info(f"Следующий слот: {next_slot_time.strftime('%Y-%m-%d %H:%M')}")

            # Вычисление времени до следующего слота
            time_to_wait = (next_slot_time - current_time).total_seconds()

            if time_to_wait > 0:
                logging.info(f"Ждем {time_to_wait / 60:.2f} минут до следующего слота.")
                time.sleep(time_to_wait)

            # После пробуждения выполняем задачи
            run_tasks()

        except Exception as e:
            logging.error(f"Ошибка в основном цикле: {e}")
            time.sleep(60)  # Ждем минуту перед повторной попыткой

if __name__ == '__main__':
    main()

